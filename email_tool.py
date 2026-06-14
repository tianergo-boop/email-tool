import os
import sys
import json
import zipfile
import smtplib
from smtplib import SMTP, SMTP_SSL
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import re
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
from email.utils import formataddr

# PDF & Excel
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook


# ══════════════════════════════════════════════════════════
#  配置文件读写
# ══════════════════════════════════════════════════════════

def get_config_path():
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "config.json")


def load_config():
    path = get_config_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_config(cfg):
    path = get_config_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ══════════════════════════════════════════════════════════
#  业务逻辑
# ══════════════════════════════════════════════════════════

def split_pdf_by_order(pdf_path, output_dir):
    """按订单号拆分 PDF，返回 {订单号: {"path": str, "pages": int}}"""
    reader = PdfReader(pdf_path)
    order_pages = {}
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        m = re.search(r'订单号\s*(\d+)', text)
        if m:
            order_pages.setdefault(m.group(1), []).append(i)

    os.makedirs(output_dir, exist_ok=True)
    result = {}
    for order_no, pages in order_pages.items():
        writer = PdfWriter()
        for p in pages:
            writer.add_page(reader.pages[p])
        out_path = os.path.join(output_dir, f"{order_no}.pdf")
        with open(out_path, "wb") as f:
            writer.write(f)
        result[order_no] = {"path": out_path, "pages": len(pages)}
    return result


def find_coa_files(order_no, coa_dir, order_batch_map):
    """根据订单号找 COA 文件，返回路径列表"""
    found = []
    if not order_no or order_no not in order_batch_map:
        return found
    for device, batch_str in order_batch_map[order_no]:
        if not device or not batch_str:
            continue
        batches = [b.strip() for b in str(batch_str).split("+") if b.strip()]
        for batch in batches:
            target = f"{device}-{batch}"
            for fname in os.listdir(coa_dir):
                fpath = os.path.join(coa_dir, fname)
                if not os.path.isfile(fpath):
                    continue
                if target.lower() in fname.lower():
                    found.append(fpath)
                    break
    seen, unique = set(), []
    for f in found:
        if f not in seen:
            seen.add(f)
            unique.append(f)
    return unique


def zip_files(file_paths, zip_path):
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in file_paths:
            zf.write(fp, os.path.basename(fp))


def read_email_data(excel_path):
    wb = load_workbook(excel_path, data_only=True)

    ws = wb["改单邮件"]
    order_email_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        order_no = row[0]
        if order_no is None:
            continue
        order_str = str(int(order_no)) if isinstance(order_no, (int, float)) else str(order_no).strip()
        if order_str in order_email_map:
            continue
        title = row[1] if row[1] else ""
        body  = row[3] if row[3] and str(row[3]).strip() != "." else ""
        order_email_map[order_str] = {
            "订单号": order_str,
            "邮件标题": str(title),
            "邮件正文": str(body) if body else "您好，请查收附件。"
        }
    email_data = list(order_email_map.values())

    ws2 = wb["订单批次"]
    order_batch = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        order_no, device, batch = row[0], row[1], row[2]
        if order_no is None or not device:
            continue
        order_str = str(int(order_no)) if isinstance(order_no, (int, float)) else str(order_no).strip()
        order_batch.setdefault(order_str, []).append((str(device), str(batch) if batch else ""))

    return email_data, order_batch


def send_email(smtp_server, smtp_port, username, password,
               recipients, subject, body, attachments=None,
               server=None):
    """发送邮件。若传入已有的 server 对象则复用连接，否则新建连接。

    返回 server 对象（保持连接），调用方负责在全部发送完毕后 quit。
    """
    # 清洗凭证：去除首尾不可见字符（全角空格、BOM 等）
    if username:
        username = username.strip().strip('\ufeff').strip('\u200b')
    if password:
        password = password.strip().strip('\ufeff').strip('\u200b')

    msg = MIMEMultipart()
    # From 头用 formataddr 编码，支持非 ASCII 显示名
    msg["From"] = formataddr((username, username))
    msg["To"]   = recipients
    msg["Subject"] = Header(subject, "utf-8")
    msg.attach(MIMEText(body, "plain", "utf-8"))

    if attachments:
        for fp in attachments:
            if not os.path.exists(fp):
                continue
            part = MIMEBase("application", "octet-stream")
            with open(fp, "rb") as f:
                part.set_payload(f.read())
            encoders.encode_base64(part)
            fname = os.path.basename(fp)
            part.add_header("Content-Disposition", "attachment",
                            filename=("utf-8", "", fname))
            msg.attach(part)

    use_ssl = (str(smtp_port) == "465")

    # 最多重试 3 次建立连接
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        try:
            if server is None:
                if use_ssl:
                    server = SMTP_SSL(smtp_server, int(smtp_port), timeout=30)
                else:
                    server = SMTP(smtp_server, int(smtp_port), timeout=30)
                    server.starttls()

            try:
                server.login(username, password)
            except UnicodeEncodeError as e:
                raise Exception(
                    "SMTP 登录失败：账号或授权码包含非英文字符。\n"
                    "请检查设置中的邮箱账号和授权码是否为英文或数字，"
                    "并确保没有多余的中文空格或不可见字符。"
                ) from e
            # login 成功即可跳出重试循环
            break
        except Exception as e:
            if server is not None:
                try:
                    server.quit()
                except Exception:
                    pass
                server = None
            if attempt < max_retries:
                import time
                time.sleep(attempt * 2)  # 2s, 4s 递增等待
            else:
                raise Exception(f"SMTP 连接失败（重试 {max_retries} 次）：{e}") from e

    try:
        rcpt_list = [r.strip() for r in recipients.split(";") if r.strip()]
        server.sendmail(username, rcpt_list, msg.as_string())
    except Exception as e:
        # sendmail 失败时连接可能已断开，废弃此 server
        try:
            server.quit()
        except Exception:
            pass
        raise Exception(f"邮件发送失败（Server not connected 或连接中断）：{e}") from e

    return server


def find_newest_file_by_keyword(search_dir, keyword):
    """在指定目录中找到包含关键词的最新文件"""
    if not os.path.isdir(search_dir):
        return None
    matches = []
    for fname in os.listdir(search_dir):
        fpath = os.path.join(search_dir, fname)
        if not os.path.isfile(fpath):
            continue
        if keyword.lower() in fname.lower():
            matches.append((fpath, os.path.getmtime(fpath)))
    if not matches:
        return None
    matches.sort(key=lambda x: x[1], reverse=True)
    return matches[0][0]


# ══════════════════════════════════════════════════════════
#  颜色主题
# ══════════════════════════════════════════════════════════

class Theme:
    # ══════════════════════════════════════════════════════
    #  主色（Google Material Design Blue）
    # ══════════════════════════════════════════════════════
    PRIMARY             = "#185FA5"
    PRIMARY_HOVER     = "#134A8C"
    PRIMARY_ACTIVE     = "#0F3A6E"
    PRIMARY_LIGHT     = "#E8F0FE"

    # ══════════════════════════════════════════════════════
    #  语义色
    # ══════════════════════════════════════════════════════
    SUCCESS            = "#3B8A3E"
    SUCCESS_BG         = "#E6F4EA"
    WARNING            = "#C47617"
    WARNING_BG         = "#FEF7E0"
    ERROR              = "#B32D2D"
    ERROR_BG           = "#FCE8E6"
    INFO               = "#185FA5"
    INFO_BG            = "#E8F0FE"

    # ══════════════════════════════════════════════════════
    #  侧边栏（深色）
    # ══════════════════════════════════════════════════════
    SIDEBAR_BG             = "#1e1e2e"
    SIDEBAR_TEXT_INACTIVE  = "#a0a0b8"
    SIDEBAR_TEXT_ACTIVE    = "#ffffff"
    SIDEBAR_ITEM_HOVER    = "#2a2a3e"
    SIDEBAR_ITEM_ACTIVE    = "#2d2d4a"
    SIDEBAR_ICON_INACTIVE = "#6c6c88"
    SIDEBAR_ICON_ACTIVE   = "#6ea8fe"
    SIDEBAR_VERSION        = "#4a4a5e"

    # ══════════════════════════════════════════════════════
    #  页面与表面
    # ══════════════════════════════════════════════════════
    BG_PAGE             = "#f8f9fa"
    BG_CARD             = "#ffffff"
    BG_SURFACE_HOVER   = "#f8f9fa"
    BORDER              = "#e8eaed"
    BORDER_HOVER       = "#c4c7cc"

    # ══════════════════════════════════════════════════════
    #  文字
    # ══════════════════════════════════════════════════════
    TEXT_PRIMARY        = "#202124"
    TEXT_SECONDARY     = "#5f6368"
    TEXT_PLACEHOLDER   = "#80868b"
    TEXT_DISABLED       = "#80868b"

    # ══════════════════════════════════════════════════════
    #  输入框
    # ══════════════════════════════════════════════════════
    INPUT_BORDER          = "#dadce0"
    INPUT_BORDER_FOCUS   = "#1a73e8"
    INPUT_BORDER_HOVER   = "#c4c7cc"
    INPUT_BG              = "#ffffff"
    INPUT_BG_DISABLED    = "#f1f3f4"
    # ════════════════════════════════════════════════════════
    #  兼容旧属性名（代码迁移过渡期）
    # ════════════════════════════════════════════════════════
    BG_INPUT              = BG_CARD
    TEXT_LABEL            = TEXT_PRIMARY
    SIDEBAR_ACTIVE_BG   = SIDEBAR_ITEM_ACTIVE
    SIDEBAR_TEXT         = SIDEBAR_TEXT_INACTIVE
    SIDEBAR_ACTIVE       = SIDEBAR_TEXT_ACTIVE
    SIDEBAR_HOVER       = SIDEBAR_ITEM_HOVER
    SIDEBAR_DIVIDER      = BORDER



# ══════════════════════════════════════════════════════════
#  GUI 主程序 v2.0
# ══════════════════════════════════════════════════════════

class EmailToolApp:

    WINDOW_W = 980
    WINDOW_H = 700
    SIDEBAR_W = 220

    def __init__(self, root):
        self.root = root
        self.root.title("邮件批量发送工具")
        self.root.geometry(f"{self.WINDOW_W}x{self.WINDOW_H}")
        self.root.minsize(860, 600)
        self.root.configure(bg=Theme.BG_PAGE)

        # ── 界面变量：改单邮件 ──
        self.order_excel_path  = tk.StringVar()
        self.order_coa_dir     = tk.StringVar()
        self.order_pdf_path    = tk.StringVar()
        self.order_output_dir  = tk.StringVar()
        self.order_smtp_server = tk.StringVar(value="smtp.exmail.qq.com")
        self.order_smtp_port   = tk.StringVar(value="465")
        self.order_email_user  = tk.StringVar()
        self.order_email_pass  = tk.StringVar()
        self.order_recipients  = tk.StringVar()

        # ── 界面变量：发货通知 ──
        self.ship_search_dir   = tk.StringVar()
        self.ship_keyword      = tk.StringVar()
        self.ship_subject_pre  = tk.StringVar(value="发货通知")
        self.ship_email_body   = tk.StringVar(value="您好，请查收附件。")
        self.ship_smtp_server  = tk.StringVar(value="smtp.exmail.qq.com")
        self.ship_smtp_port    = tk.StringVar(value="465")
        self.ship_email_user   = tk.StringVar()
        self.ship_email_pass   = tk.StringVar()
        self.ship_recipients   = tk.StringVar()

        # ── 通用设置 ──
        self.confirm_before_send = tk.BooleanVar(value=True)

        # ── 运行状态 ──
        self._running = False

        self._load_saved_config()
        self._setup_styles()
        self._build_ui()

    # ════════════════════════════════════════════════════════
    #  配置持久化
    # ════════════════════════════════════════════════════════

    def _load_saved_config(self):
        cfg = load_config()
        # 兼容旧版配置
        o = cfg.get("order_email", {})
        if not o and cfg.get("excel_path"):  # 旧版迁移
            o = {k: cfg[k] for k in [
                "excel_path", "coa_dir", "pdf_path", "output_dir",
                "smtp_server", "smtp_port", "email_user", "email_pass", "recipients"
            ] if k in cfg}
            o.setdefault("smtp_server", "smtp.exmail.qq.com")
            o.setdefault("smtp_port", "465")

        if o.get("excel_path"):  self.order_excel_path.set(o["excel_path"])
        if o.get("coa_dir"):    self.order_coa_dir.set(o["coa_dir"])
        if o.get("pdf_path"):   self.order_pdf_path.set(o["pdf_path"])
        if o.get("output_dir"): self.order_output_dir.set(o["output_dir"])
        if o.get("smtp_server"): self.order_smtp_server.set(o["smtp_server"])
        if o.get("smtp_port"):   self.order_smtp_port.set(o["smtp_port"])
        if o.get("email_user"):  self.order_email_user.set(o["email_user"])
        if o.get("email_pass"):  self.order_email_pass.set(o["email_pass"])
        if o.get("recipients"):  self.order_recipients.set(o["recipients"])

        s = cfg.get("shipping_notify", {})
        if s.get("search_dir"):   self.ship_search_dir.set(s["search_dir"])
        if s.get("keyword"):      self.ship_keyword.set(s["keyword"])
        if s.get("subject_prefix"): self.ship_subject_pre.set(s["subject_prefix"])
        if s.get("email_body"):   self.ship_email_body.set(s["email_body"])
        if s.get("smtp_server"):  self.ship_smtp_server.set(s["smtp_server"])
        if s.get("smtp_port"):    self.ship_smtp_port.set(s["smtp_port"])
        if s.get("email_user"):   self.ship_email_user.set(s["email_user"])
        if s.get("email_pass"):   self.ship_email_pass.set(s["email_pass"])
        if s.get("recipients"):   self.ship_recipients.set(s["recipients"])

        g = cfg.get("general", {})
        if "confirm_before_send" in g:
            self.confirm_before_send.set(g["confirm_before_send"])

    def _save_current_config(self):
        cfg = {
            "order_email": {
                "excel_path":  self.order_excel_path.get(),
                "coa_dir":     self.order_coa_dir.get(),
                "pdf_path":    self.order_pdf_path.get(),
                "output_dir":  self.order_output_dir.get(),
                "smtp_server": self.order_smtp_server.get(),
                "smtp_port":   self.order_smtp_port.get(),
                "email_user":  self.order_email_user.get(),
                "email_pass":  self.order_email_pass.get(),
                "recipients":  self.order_recipients.get(),
            },
            "shipping_notify": {
                "search_dir":    self.ship_search_dir.get(),
                "keyword":       self.ship_keyword.get(),
                "subject_prefix": self.ship_subject_pre.get(),
                "email_body":    self.ship_email_body.get(),
                "smtp_server":   self.ship_smtp_server.get(),
                "smtp_port":     self.ship_smtp_port.get(),
                "email_user":    self.ship_email_user.get(),
                "email_pass":    self.ship_email_pass.get(),
                "recipients":    self.ship_recipients.get(),
            },
            "general": {
                "confirm_before_send": self.confirm_before_send.get(),
            }
        }
        save_config(cfg)

    # ════════════════════════════════════════════════════════
    #  样式
    # ════════════════════════════════════════════════════════

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        # ── 卡片标题 ──
        style.configure("Card.TLabelframe",
                        background=Theme.BG_CARD,
                        bordercolor=Theme.BORDER,
                        relief="solid", borderwidth=1)
        style.configure("Card.TLabelframe.Label",
                        background=Theme.BG_CARD,
                        foreground=Theme.TEXT_PRIMARY,
                        font=("Microsoft YaHei UI", 11, "bold"),
                        padding=(16, 6, 16, 4))

        # ── 输入框 ──
        style.configure("Input.TEntry",
                        fieldbackground=Theme.INPUT_BG,
                        bordercolor=Theme.INPUT_BORDER,
                        focuscolor=Theme.INPUT_BORDER_FOCUS,
                        padding=(8, 6),
                        font=("Microsoft YaHei UI", 10))

        # ── 主按钮（Primary）──
        style.configure("Primary.TButton",
                        background=Theme.PRIMARY,
                        foreground="white",
                        font=("Microsoft YaHei UI", 10, "bold"),
                        padding=(24, 10), borderwidth=0)
        style.map("Primary.TButton",
                  background=[("active", Theme.PRIMARY_HOVER),
                             ("pressed", Theme.PRIMARY_ACTIVE)],
                  relief=[("pressed", "sunken")])

        # ── 次按钮（Secondary）──
        style.configure("Secondary.TButton",
                        background=Theme.BG_CARD,
                        foreground=Theme.PRIMARY,
                        font=("Microsoft YaHei UI", 9),
                        padding=(14, 6), borderwidth=1,
                        bordercolor=Theme.PRIMARY)
        style.map("Secondary.TButton",
                  background=[("active", Theme.PRIMARY_LIGHT)],
                  foreground=[("active", Theme.PRIMARY_HOVER)],
                  bordercolor=[("active", Theme.PRIMARY_HOVER)])

        # ── 进度条 ──
        style.configure("Horizontal.TProgressbar",
                        troughcolor=Theme.BORDER,
                        background=Theme.PRIMARY,
                        borderwidth=0, thickness=6)

        # ── 复选框 ──
        style.configure("TCheckbutton",
                        background=Theme.BG_CARD,
                        font=("Microsoft YaHei UI", 9),
                        foreground=Theme.TEXT_PRIMARY)
        style.map("TCheckbutton",
                  background=[("active", Theme.BG_CARD)])

    # ════════════════════════════════════════════════════════
    #  主界面搭建
    # ════════════════════════════════════════════════════════

    def _build_ui(self):
        # 侧边栏
        self.sidebar = tk.Frame(self.root, bg=Theme.SIDEBAR_BG,
                                width=self.SIDEBAR_W)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar()

        # 主内容区
        self.main_area = tk.Frame(self.root, bg=Theme.BG_PAGE)
        self.main_area.pack(side="left", fill="both", expand=True)

        # 页面容器
        self.pages = {}
        self._build_order_email_page()
        self._build_shipping_notify_page()
        self._build_settings_page()

        # 默认显示改单邮件页面
        self.current_page = None
        self._switch_page("order_email")

    # ───── 侧边栏 ─────

    def _build_sidebar(self):
        sb = self.sidebar

        # Logo 区域
        logo_frame = tk.Frame(sb, bg=Theme.SIDEBAR_BG)
        logo_frame.pack(fill="x", padx=16, pady=(20, 4))
        tk.Label(logo_frame, text="✉", font=("Segoe UI Emoji", 20),
                 bg=Theme.SIDEBAR_BG, fg=Theme.PRIMARY).pack(side="left")
        tk.Label(logo_frame, text="邮件工具",
                 font=("Microsoft YaHei UI", 13, "bold"),
                 bg=Theme.SIDEBAR_BG, fg="#FFFFFF").pack(side="left", padx=(8, 0))

        tk.Label(sb, text="v2.2",
                 font=("Microsoft YaHei UI", 8),
                 bg=Theme.SIDEBAR_BG, fg=Theme.SIDEBAR_VERSION).pack(anchor="w", padx=20)

        # 分割线
        tk.Frame(sb, bg=Theme.BORDER, height=1).pack(fill="x", padx=16, pady=(16, 8))

        # 导航标题
        tk.Label(sb, text="功  能",
                 font=("Microsoft YaHei UI", 9),
                 bg=Theme.SIDEBAR_BG, fg=Theme.SIDEBAR_TEXT_INACTIVE).pack(anchor="w", padx=20, pady=(4, 6))

        # 导航按钮
        self.nav_buttons = {}

        btn_order = tk.Frame(sb, bg=Theme.SIDEBAR_BG, cursor="hand2")
        btn_order.pack(fill="x", padx=8, pady=2)
        btn_order.bind("<Button-1>", lambda e: self._switch_page("order_email"))
        self._nav_order_indicator = tk.Frame(btn_order, bg=Theme.SIDEBAR_BG, width=3)
        self._nav_order_indicator.pack(side="left", fill="y", pady=6)
        self._nav_order_inner = tk.Frame(btn_order, bg=Theme.SIDEBAR_BG)
        self._nav_order_inner.pack(side="left", fill="x", expand=True, padx=(8, 8), pady=8)
        self._nav_order_label = tk.Label(self._nav_order_inner, text="📋  改单邮件",
                                          font=("Microsoft YaHei UI", 10),
                                          bg=Theme.SIDEBAR_BG, fg=Theme.SIDEBAR_TEXT_INACTIVE)
        self._nav_order_label.pack(anchor="w")
        for w in [btn_order, self._nav_order_inner, self._nav_order_label]:
            w.bind("<Button-1>", lambda e: self._switch_page("order_email"))
            w.bind("<Enter>", lambda e, f=btn_order: f.configure(bg=Theme.SIDEBAR_HOVER) if self.current_page != "order_email" else None)
            w.bind("<Leave>", lambda e, f=btn_order: f.configure(bg=Theme.SIDEBAR_BG) if self.current_page != "order_email" else None)
        self.nav_buttons["order_email"] = (btn_order, self._nav_order_indicator, self._nav_order_label, self._nav_order_inner)

        btn_ship = tk.Frame(sb, bg=Theme.SIDEBAR_BG, cursor="hand2")
        btn_ship.pack(fill="x", padx=8, pady=2)
        btn_ship.bind("<Button-1>", lambda e: self._switch_page("shipping_notify"))
        self._nav_ship_indicator = tk.Frame(btn_ship, bg=Theme.SIDEBAR_BG, width=3)
        self._nav_ship_indicator.pack(side="left", fill="y", pady=6)
        self._nav_ship_inner = tk.Frame(btn_ship, bg=Theme.SIDEBAR_BG)
        self._nav_ship_inner.pack(side="left", fill="x", expand=True, padx=(8, 8), pady=8)
        self._nav_ship_label = tk.Label(self._nav_ship_inner, text="🚚  发货通知",
                                          font=("Microsoft YaHei UI", 10),
                                          bg=Theme.SIDEBAR_BG, fg=Theme.SIDEBAR_TEXT_INACTIVE)
        self._nav_ship_label.pack(anchor="w")
        for w in [btn_ship, self._nav_ship_inner, self._nav_ship_label]:
            w.bind("<Button-1>", lambda e: self._switch_page("shipping_notify"))
            w.bind("<Enter>", lambda e, f=btn_ship: f.configure(bg=Theme.SIDEBAR_HOVER) if self.current_page != "shipping_notify" else None)
            w.bind("<Leave>", lambda e, f=btn_ship: f.configure(bg=Theme.SIDEBAR_BG) if self.current_page != "shipping_notify" else None)
        self.nav_buttons["shipping_notify"] = (btn_ship, self._nav_ship_indicator, self._nav_ship_label, self._nav_ship_inner)

        # ── 配箱按钮（暂未启用，待优化后开放） ──
        # btn_alloc = tk.Frame(sb, bg=Theme.SIDEBAR_BG, cursor="hand2")
        # btn_alloc.pack(fill="x", padx=8, pady=2)
        # btn_alloc.bind("<Button-1>", lambda e: self._switch_page("allocation"))
        # self._nav_alloc_indicator = tk.Frame(btn_alloc, bg=Theme.SIDEBAR_BG, width=3)
        # self._nav_alloc_indicator.pack(side="left", fill="y", pady=6)
        # self._nav_alloc_inner = tk.Frame(btn_alloc, bg=Theme.SIDEBAR_BG)
        # self._nav_alloc_inner.pack(side="left", fill="x", expand=True, padx=(8, 8), pady=8)
        # self._nav_alloc_label = tk.Label(self._nav_alloc_inner, text="📦  配箱工具",
        #                                     font=("Microsoft YaHei UI", 10),
        #                                     bg=Theme.SIDEBAR_BG, fg=Theme.SIDEBAR_TEXT_INACTIVE)
        # self._nav_alloc_label.pack(anchor="w")
        # for w in [btn_alloc, self._nav_alloc_inner, self._nav_alloc_label]:
        #     w.bind("<Button-1>", lambda e: self._switch_page("allocation"))
        #     w.bind("<Enter>", lambda e, f=btn_alloc: f.configure(bg=Theme.SIDEBAR_HOVER) if self.current_page != "allocation" else None)
        #     w.bind("<Leave>", lambda e, f=btn_alloc: f.configure(bg=Theme.SIDEBAR_BG) if self.current_page != "allocation" else None)
        # self.nav_buttons["allocation"] = (btn_alloc, self._nav_alloc_indicator, self._nav_alloc_label, self._nav_alloc_inner)

        # 底部设置按钮
        tk.Frame(sb, bg=Theme.SIDEBAR_DIVIDER, height=1).pack(fill="x", padx=16, pady=(8, 8), side="bottom", before=None)

        # 弹性空间推到底部
        tk.Frame(sb, bg=Theme.SIDEBAR_BG).pack(fill="both", expand=True)

        btn_settings = tk.Frame(sb, bg=Theme.SIDEBAR_BG, cursor="hand2")
        btn_settings.pack(fill="x", padx=8, pady=(2, 12), side="bottom")
        self._nav_settings_indicator = tk.Frame(btn_settings, bg=Theme.SIDEBAR_BG, width=3)
        self._nav_settings_indicator.pack(side="left", fill="y", pady=6)
        self._nav_settings_inner = tk.Frame(btn_settings, bg=Theme.SIDEBAR_BG)
        self._nav_settings_inner.pack(side="left", fill="x", expand=True, padx=(8, 8), pady=8)
        self._nav_settings_label = tk.Label(self._nav_settings_inner, text="⚙️  设置",
                                              font=("Microsoft YaHei UI", 10),
                                              bg=Theme.SIDEBAR_BG, fg=Theme.SIDEBAR_TEXT)
        self._nav_settings_label.pack(anchor="w")
        for w in [btn_settings, self._nav_settings_inner, self._nav_settings_label]:
            w.bind("<Button-1>", lambda e: self._switch_page("settings"))
            w.bind("<Enter>", lambda e, f=btn_settings: f.configure(bg=Theme.SIDEBAR_HOVER) if self.current_page != "settings" else None)
            w.bind("<Leave>", lambda e, f=btn_settings: f.configure(bg=Theme.SIDEBAR_BG) if self.current_page != "settings" else None)
        self.nav_buttons["settings"] = (btn_settings, self._nav_settings_indicator, self._nav_settings_label, self._nav_settings_inner)

    def _switch_page(self, page_name):
        # 隐藏当前页面
        if self.current_page and self.current_page in self.pages:
            self.pages[self.current_page].pack_forget()

        # 重置旧导航高亮
        for name, (btn, indicator, label, inner) in self.nav_buttons.items():
            indicator.configure(bg=Theme.SIDEBAR_BG)
            btn.configure(bg=Theme.SIDEBAR_BG)
            inner.configure(bg=Theme.SIDEBAR_BG)
            label.configure(bg=Theme.SIDEBAR_BG, fg=Theme.SIDEBAR_TEXT)

        # 高亮新导航
        if page_name in self.nav_buttons:
            btn, indicator, label, inner = self.nav_buttons[page_name]
            indicator.configure(bg=Theme.PRIMARY)
            btn.configure(bg=Theme.SIDEBAR_HOVER)
            inner.configure(bg=Theme.SIDEBAR_HOVER)
            label.configure(bg=Theme.SIDEBAR_HOVER, fg=Theme.SIDEBAR_ACTIVE)

        self.current_page = page_name
        self.pages[page_name].pack(in_=self.main_area, fill="both", expand=True)

        # 切换到功能页面时刷新状态指示器
        if page_name == "order_email":
            self._refresh_order_status()
        elif page_name == "shipping_notify":
            self._refresh_ship_status()
        elif page_name == "allocation":
            pass  # 配箱功能暂未启用

    # ════════════════════════════════════════════════════════
    #  改单邮件页面
    # ════════════════════════════════════════════════════════

    def _build_order_email_page(self):
        page = tk.Frame(self.main_area, bg=Theme.BG_PAGE)

        # 标题
        header = tk.Frame(page, bg=Theme.BG_PAGE)
        header.pack(fill="x", padx=28, pady=(20, 12))
        tk.Label(header, text="改单邮件",
                 font=("Microsoft YaHei UI", 18, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_PAGE).pack(side="left")
        tk.Label(header, text="按订单号拆分PDF并发送邮件",
                 font=("Microsoft YaHei UI", 9),
                 fg=Theme.TEXT_SECONDARY, bg=Theme.BG_PAGE).pack(side="left", padx=(12, 0), pady=(6, 0))

        # 配置状态卡片
        status_card = tk.Frame(page, bg=Theme.BG_CARD,
                               highlightbackground=Theme.BORDER,
                               highlightthickness=1)
        status_card.pack(fill="x", padx=28, pady=(0, 16))

        status_header = tk.Frame(status_card, bg=Theme.BG_CARD)
        status_header.pack(fill="x", padx=20, pady=(14, 8))
        tk.Label(status_header, text="配置状态",
                 font=("Microsoft YaHei UI", 10, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        self.order_status_frame = tk.Frame(status_card, bg=Theme.BG_CARD)
        self.order_status_frame.pack(fill="x", padx=20, pady=(0, 14))

        self.order_status_labels = {}
        status_items = [
            ("excel", "Excel 文件"),
            ("coa",   "COA 文件夹"),
            ("pdf",   "PDF 文件"),
            ("output","输出目录"),
            ("smtp",  "邮箱配置"),
            ("recipients", "收件人"),
        ]
        for i, (key, label) in enumerate(status_items):
            row = tk.Frame(self.order_status_frame, bg=Theme.BG_CARD)
            row.pack(fill="x", pady=3)
            lbl = tk.Label(row, text="❌  " + label,
                          font=("Microsoft YaHei UI", 9),
                          fg=Theme.TEXT_SECONDARY, bg=Theme.BG_CARD)
            lbl.pack(side="left")
            self.order_status_labels[key] = lbl

        # 操作区
        action_frame = tk.Frame(page, bg=Theme.BG_PAGE)
        action_frame.pack(fill="x", padx=28, pady=(0, 12))

        self.order_btn_run = ttk.Button(action_frame, text="▶  开始执行",
                                         style="Primary.TButton",
                                         command=self._on_order_execute)
        self.order_btn_run.pack(side="left")

        self.order_progress = ttk.Progressbar(action_frame, mode="determinate",
                                               maximum=100,
                                               style="Horizontal.TProgressbar",
                                               length=320)
        self.order_progress.pack(side="left", padx=(20, 0), fill="x", expand=True)

        self.order_progress_label = tk.Label(action_frame, text="",
                                              font=("Microsoft YaHei UI", 9),
                                              fg=Theme.TEXT_SECONDARY, bg=Theme.BG_PAGE)
        self.order_progress_label.pack(side="left", padx=(8, 0))

        # 日志区
        log_card = tk.Frame(page, bg=Theme.BG_CARD,
                           highlightbackground=Theme.BORDER,
                           highlightthickness=1)
        log_card.pack(fill="both", expand=True, padx=28, pady=(0, 16))

        log_header = tk.Frame(log_card, bg=Theme.BG_CARD)
        log_header.pack(fill="x", padx=16, pady=(10, 4))
        tk.Label(log_header, text="执行记录",
                 font=("Microsoft YaHei UI", 10, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        self.order_log_text = tk.Text(
            log_card, height=10, state="disabled",
            font=("Consolas", 9), wrap="word",
            bg="#FAFBFC", fg=Theme.TEXT_PRIMARY,
            borderwidth=0, highlightthickness=0,
            insertbackground=Theme.TEXT_PRIMARY,
            selectbackground=Theme.PRIMARY,
            selectforeground="white",
            padx=12, pady=8
        )
        log_scroll = ttk.Scrollbar(log_card, command=self.order_log_text.yview)
        self.order_log_text.configure(yscrollcommand=log_scroll.set)
        log_scroll.pack(side="right", fill="y", padx=(0, 4), pady=(0, 8))
        self.order_log_text.pack(fill="both", expand=True, padx=(12, 0), pady=(0, 8))

        self.order_log_text.tag_configure("red",    foreground=Theme.ERROR)
        self.order_log_text.tag_configure("green",  foreground=Theme.SUCCESS)
        self.order_log_text.tag_configure("warn",   foreground=Theme.WARNING)
        self.order_log_text.tag_configure("header", foreground=Theme.TEXT_SECONDARY)

        self.pages["order_email"] = page

    def _refresh_order_status(self):
        checks = {
            "excel":      bool(self.order_excel_path.get() and os.path.isfile(self.order_excel_path.get())),
            "coa":        bool(self.order_coa_dir.get() and os.path.isdir(self.order_coa_dir.get())),
            "pdf":        bool(self.order_pdf_path.get() and os.path.isfile(self.order_pdf_path.get())),
            "output":     bool(self.order_output_dir.get()),
            "smtp":       bool(self.order_email_user.get() and self.order_email_pass.get()),
            "recipients": bool(self.order_recipients.get()),
        }
        for key, ok in checks.items():
            lbl = self.order_status_labels[key]
            name = lbl.cget("text").split("  ", 1)[-1] if "  " in lbl.cget("text") else lbl.cget("text")[3:]
            if ok:
                lbl.configure(text=f"✅  {name}", fg=Theme.SUCCESS)
            else:
                lbl.configure(text=f"❌  {name}", fg=Theme.TEXT_SECONDARY)

    # ════════════════════════════════════════════════════════
    #  发货通知页面
    # ════════════════════════════════════════════════════════

    def _build_shipping_notify_page(self):
        page = tk.Frame(self.main_area, bg=Theme.BG_PAGE)

        # 标题
        header = tk.Frame(page, bg=Theme.BG_PAGE)
        header.pack(fill="x", padx=28, pady=(20, 12))
        tk.Label(header, text="发货通知",
                 font=("Microsoft YaHei UI", 18, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_PAGE).pack(side="left")
        tk.Label(header, text="查找文件并发送发货通知邮件",
                 font=("Microsoft YaHei UI", 9),
                 fg=Theme.TEXT_SECONDARY, bg=Theme.BG_PAGE).pack(side="left", padx=(12, 0), pady=(6, 0))

        # 配置状态卡片
        status_card = tk.Frame(page, bg=Theme.BG_CARD,
                               highlightbackground=Theme.BORDER,
                               highlightthickness=1)
        status_card.pack(fill="x", padx=28, pady=(0, 16))

        status_header = tk.Frame(status_card, bg=Theme.BG_CARD)
        status_header.pack(fill="x", padx=20, pady=(14, 8))
        tk.Label(status_header, text="配置状态",
                 font=("Microsoft YaHei UI", 10, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        self.ship_status_frame = tk.Frame(status_card, bg=Theme.BG_CARD)
        self.ship_status_frame.pack(fill="x", padx=20, pady=(0, 14))

        self.ship_status_labels = {}
        ship_status_items = [
            ("search_dir", "搜索文件夹"),
            ("keyword",    "文件特征值"),
            ("subject",    "邮件标题前缀"),
            ("smtp",       "邮箱配置"),
            ("recipients", "收件人"),
        ]
        for i, (key, label) in enumerate(ship_status_items):
            row = tk.Frame(self.ship_status_frame, bg=Theme.BG_CARD)
            row.pack(fill="x", pady=3)
            lbl = tk.Label(row, text=f"❌  {label}",
                          font=("Microsoft YaHei UI", 9),
                          fg=Theme.TEXT_SECONDARY, bg=Theme.BG_CARD)
            lbl.pack(side="left")
            self.ship_status_labels[key] = lbl

        # 操作区
        action_frame = tk.Frame(page, bg=Theme.BG_PAGE)
        action_frame.pack(fill="x", padx=28, pady=(0, 12))

        self.ship_btn_run = ttk.Button(action_frame, text="▶  开始执行",
                                        style="Primary.TButton",
                                        command=self._on_ship_execute)
        self.ship_btn_run.pack(side="left")

        self.ship_progress = ttk.Progressbar(action_frame, mode="determinate",
                                              maximum=1,
                                              style="Horizontal.TProgressbar",
                                              length=320)
        self.ship_progress.pack(side="left", padx=(20, 0), fill="x", expand=True)

        self.ship_progress_label = tk.Label(action_frame, text="",
                                             font=("Microsoft YaHei UI", 9),
                                             fg=Theme.TEXT_SECONDARY, bg=Theme.BG_PAGE)
        self.ship_progress_label.pack(side="left", padx=(8, 0))

        # 日志区
        log_card = tk.Frame(page, bg=Theme.BG_CARD,
                           highlightbackground=Theme.BORDER,
                           highlightthickness=1)
        log_card.pack(fill="both", expand=True, padx=28, pady=(0, 16))

        log_header = tk.Frame(log_card, bg=Theme.BG_CARD)
        log_header.pack(fill="x", padx=16, pady=(10, 4))
        tk.Label(log_header, text="执行记录",
                 font=("Microsoft YaHei UI", 10, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        self.ship_log_text = tk.Text(
            log_card, height=10, state="disabled",
            font=("Consolas", 9), wrap="word",
            bg="#FAFBFC", fg=Theme.TEXT_PRIMARY,
            borderwidth=0, highlightthickness=0,
            insertbackground=Theme.TEXT_PRIMARY,
            selectbackground=Theme.PRIMARY,
            selectforeground="white",
            padx=12, pady=8
        )
        log_scroll = ttk.Scrollbar(log_card, command=self.ship_log_text.yview)
        self.ship_log_text.configure(yscrollcommand=log_scroll.set)
        log_scroll.pack(side="right", fill="y", padx=(0, 4), pady=(0, 8))
        self.ship_log_text.pack(fill="both", expand=True, padx=(12, 0), pady=(0, 8))

        self.ship_log_text.tag_configure("red",    foreground=Theme.ERROR)
        self.ship_log_text.tag_configure("green",  foreground=Theme.SUCCESS)
        self.ship_log_text.tag_configure("warn",   foreground=Theme.WARNING)
        self.ship_log_text.tag_configure("header", foreground=Theme.TEXT_SECONDARY)

        self.pages["shipping_notify"] = page

    def _refresh_ship_status(self):
        checks = {
            "search_dir": bool(self.ship_search_dir.get() and os.path.isdir(self.ship_search_dir.get())),
            "keyword":    bool(self.ship_keyword.get().strip()),
            "subject":    bool(self.ship_subject_pre.get().strip()),
            "smtp":       bool(self.ship_email_user.get() and self.ship_email_pass.get()),
            "recipients": bool(self.ship_recipients.get()),
        }
        for key, ok in checks.items():
            lbl = self.ship_status_labels[key]
            name = lbl.cget("text")[3:]
            if ok:
                lbl.configure(text=f"✅  {name}", fg=Theme.SUCCESS)
            else:
                lbl.configure(text=f"❌  {name}", fg=Theme.TEXT_SECONDARY)

    # ════════════════════════════════════════════════════════
    #  设置页面
    # ════════════════════════════════════════════════════════

    def _build_settings_page(self):
        page = tk.Frame(self.main_area, bg=Theme.BG_PAGE)

        # 标题
        header = tk.Frame(page, bg=Theme.BG_PAGE)
        header.pack(fill="x", padx=28, pady=(20, 12))
        tk.Label(header, text="设置",
                 font=("Microsoft YaHei UI", 18, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_PAGE).pack(side="left")

        # 可滚动区域
        canvas = tk.Canvas(page, bg=Theme.BG_PAGE, highlightthickness=0)
        scrollbar = ttk.Scrollbar(page, orient="vertical", command=canvas.yview)
        self.settings_scroll_frame = tk.Frame(canvas, bg=Theme.BG_PAGE)

        self.settings_scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.settings_scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True, padx=(28, 0), pady=(0, 16))
        scrollbar.pack(side="right", fill="y", padx=(0, 12), pady=(0, 16))

        # 鼠标滚轮支持
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        sf = self.settings_scroll_frame
        sf.configure(width=800)

        # ── 改单邮件设置 ──
        self._build_settings_section(sf, "改单邮件", [
            ("Excel 文件",  self.order_excel_path,  self._select_order_excel,  "file"),
            ("COA 文件夹",  self.order_coa_dir,     self._select_order_coa,    "dir"),
            ("PDF 文件",    self.order_pdf_path,     self._select_order_pdf,    "file"),
            ("输出目录",    self.order_output_dir,   self._select_order_output, "dir"),
        ])

        self._build_smtp_settings(sf, "改单邮件 · 邮箱配置",
                                  self.order_smtp_server, self.order_smtp_port,
                                  self.order_email_user, self.order_email_pass)

        self._build_recipient_settings(sf, "改单邮件 · 收件人",
                                       self.order_recipients)

        # ── 发货通知设置 ──
        self._build_settings_section(sf, "发货通知", [
            ("搜索文件夹",    self.ship_search_dir,   self._select_ship_dir,   "dir"),
            ("文件特征值",    self.ship_keyword,       None,                    "entry"),
            ("邮件标题前缀", self.ship_subject_pre,    None,                    "entry"),
            ("邮件正文",     self.ship_email_body,     None,                    "entry"),
        ])

        self._build_smtp_settings(sf, "发货通知 · 邮箱配置",
                                  self.ship_smtp_server, self.ship_smtp_port,
                                  self.ship_email_user, self.ship_email_pass)

        self._build_recipient_settings(sf, "发货通知 · 收件人",
                                       self.ship_recipients)

        # ── 通用设置 ──
        general_card = tk.Frame(sf, bg=Theme.BG_CARD,
                                highlightbackground=Theme.BORDER,
                                highlightthickness=1)
        general_card.pack(fill="x", pady=(12, 20), padx=4)

        gen_header = tk.Frame(general_card, bg=Theme.BG_CARD)
        gen_header.pack(fill="x", padx=20, pady=(14, 8))
        tk.Label(gen_header, text="通用设置",
                 font=("Microsoft YaHei UI", 11, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        gen_inner = tk.Frame(general_card, bg=Theme.BG_CARD)
        gen_inner.pack(fill="x", padx=20, pady=(0, 14))

        cb = tk.Checkbutton(gen_inner, text="发送前显示确认预览",
                            variable=self.confirm_before_send,
                            font=("Microsoft YaHei UI", 9),
                            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
                            selectcolor=Theme.BG_INPUT,
                            activebackground=Theme.BG_CARD,
                            activeforeground=Theme.TEXT_PRIMARY,
                            command=self._save_current_config)
        cb.pack(anchor="w", pady=4)

        tk.Label(gen_inner, text="开启后，执行前会显示待发送邮件的预览，确认后才会发送",
                 font=("Microsoft YaHei UI", 8),
                 fg=Theme.TEXT_SECONDARY, bg=Theme.BG_CARD).pack(anchor="w", pady=(0, 4))

        self.pages["settings"] = page

    def _build_settings_section(self, parent, title, items):
        """构建一个设置区段卡片"""
        card = tk.Frame(parent, bg=Theme.BG_CARD,
                        highlightbackground=Theme.BORDER,
                        highlightthickness=1)
        card.pack(fill="x", pady=(12, 0), padx=4)

        # 卡片标题
        header = tk.Frame(card, bg=Theme.BG_CARD)
        header.pack(fill="x", padx=20, pady=(14, 8))
        tk.Label(header, text=title,
                 font=("Microsoft YaHei UI", 11, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        inner = tk.Frame(card, bg=Theme.BG_CARD)
        inner.pack(fill="x", padx=20, pady=(0, 14))

        for i, (label, var, cmd, entry_type) in enumerate(items):
            row = tk.Frame(inner, bg=Theme.BG_CARD)
            row.pack(fill="x", pady=5)

            tk.Label(row, text=label, font=("Microsoft YaHei UI", 9),
                     fg=Theme.TEXT_LABEL, bg=Theme.BG_CARD, width=12, anchor="w"
                     ).pack(side="left")

            entry = ttk.Entry(row, textvariable=var, style="Input.TEntry", width=50)
            entry.pack(side="left", padx=(8, 0), fill="x", expand=True)

            if cmd:
                ttk.Button(row, text="浏览…", style="Secondary.TButton",
                           command=cmd).pack(side="left", padx=(8, 0))

    def _build_smtp_settings(self, parent, title, server_var, port_var, user_var, pass_var):
        card = tk.Frame(parent, bg=Theme.BG_CARD,
                        highlightbackground=Theme.BORDER,
                        highlightthickness=1)
        card.pack(fill="x", pady=(12, 0), padx=4)

        header = tk.Frame(card, bg=Theme.BG_CARD)
        header.pack(fill="x", padx=20, pady=(14, 8))
        tk.Label(header, text=title,
                 font=("Microsoft YaHei UI", 11, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        inner = tk.Frame(card, bg=Theme.BG_CARD)
        inner.pack(fill="x", padx=20, pady=(0, 14))

        # Row 1: SMTP + 端口
        r1 = tk.Frame(inner, bg=Theme.BG_CARD)
        r1.pack(fill="x", pady=5)
        tk.Label(r1, text="SMTP 服务器", font=("Microsoft YaHei UI", 9),
                 fg=Theme.TEXT_LABEL, bg=Theme.BG_CARD, width=12, anchor="w"
                 ).pack(side="left")
        ttk.Entry(r1, textvariable=server_var, style="Input.TEntry", width=30
                  ).pack(side="left", padx=(8, 16))
        tk.Label(r1, text="端口", font=("Microsoft YaHei UI", 9),
                 fg=Theme.TEXT_LABEL, bg=Theme.BG_CARD
                 ).pack(side="left")
        ttk.Entry(r1, textvariable=port_var, style="Input.TEntry", width=8
                  ).pack(side="left", padx=(8, 0))

        # Row 2: 账号 + 授权码
        r2 = tk.Frame(inner, bg=Theme.BG_CARD)
        r2.pack(fill="x", pady=5)
        tk.Label(r2, text="邮箱账号", font=("Microsoft YaHei UI", 9),
                 fg=Theme.TEXT_LABEL, bg=Theme.BG_CARD, width=12, anchor="w"
                 ).pack(side="left")
        ttk.Entry(r2, textvariable=user_var, style="Input.TEntry", width=30
                  ).pack(side="left", padx=(8, 16))
        tk.Label(r2, text="授权码", font=("Microsoft YaHei UI", 9),
                 fg=Theme.TEXT_LABEL, bg=Theme.BG_CARD
                 ).pack(side="left")
        ttk.Entry(r2, textvariable=pass_var, style="Input.TEntry", width=18, show="•"
                  ).pack(side="left", padx=(8, 0))

    def _build_recipient_settings(self, parent, title, recipients_var):
        card = tk.Frame(parent, bg=Theme.BG_CARD,
                        highlightbackground=Theme.BORDER,
                        highlightthickness=1)
        card.pack(fill="x", pady=(12, 0), padx=4)

        header = tk.Frame(card, bg=Theme.BG_CARD)
        header.pack(fill="x", padx=20, pady=(14, 8))
        tk.Label(header, text=title,
                 font=("Microsoft YaHei UI", 11, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        inner = tk.Frame(card, bg=Theme.BG_CARD)
        inner.pack(fill="x", padx=20, pady=(0, 14))

        ttk.Entry(inner, textvariable=recipients_var, style="Input.TEntry", width=70
                  ).pack(fill="x", pady=4)
        tk.Label(inner, text="多个收件人用英文分号 ; 分隔",
                 font=("Microsoft YaHei UI", 8),
                 fg=Theme.TEXT_SECONDARY, bg=Theme.BG_CARD, anchor="w"
                 ).pack(fill="x")

    # ════════════════════════════════════════════════════
    #  配箱工具页面
    # ════════════════════════════════════════════════════

    def _build_allocation_page(self):
        """构建配箱工具页面"""
        page = tk.Frame(self.main_area, bg=Theme.BG_PAGE)

        # 标题区
        header = tk.Frame(page, bg=Theme.BG_PAGE)
        header.pack(fill="x", padx=28, pady=(20, 12))
        tk.Label(header, text="配箱工具",
                 font=("Microsoft YaHei UI", 18, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_PAGE).pack(side="left")
        tk.Label(header, text="导入订单 → 配箱 → 导出 → 发邮件",
                 font=("Microsoft YaHei UI", 9),
                 fg=Theme.TEXT_SECONDARY, bg=Theme.BG_PAGE).pack(side="left", padx=(12, 0), pady=(6, 0))

        # 文件选择区
        file_card = tk.Frame(page, bg=Theme.BG_CARD,
                              highlightbackground=Theme.BORDER, highlightthickness=1)
        file_card.pack(fill="x", padx=28, pady=(0, 12))

        file_inner = tk.Frame(file_card, bg=Theme.BG_CARD)
        file_inner.pack(fill="x", padx=20, pady=14)

        tk.Label(file_inner, text="配箱表文件：",
                 font=("Microsoft YaHei UI", 9),
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side="left")

        self.alloc_file_var = tk.StringVar()
        entry = ttk.Entry(file_inner, textvariable=self.alloc_file_var,
                          style="Input.TEntry", width=50)
        entry.pack(side="left", padx=(8, 0))

        def _select_alloc_file():
            p = filedialog.askopenfilename(
                title="选择配箱表",
                filetypes=[("Excel 文件", "*.xlsm *.xlsx")])
            if p:
                self.alloc_file_var.set(p)
                self._load_allocation_data()

        ttk.Button(file_inner, text="浏览…", style="Secondary.TButton",
                    command=_select_alloc_file).pack(side="left", padx=(8, 0))
        ttk.Button(file_inner, text="🔄 刷新数据", style="Secondary.TButton",
                    command=lambda: self._load_allocation_data()).pack(side="left", padx=(8, 0))

        # 统计栏
        stats_card = tk.Frame(page, bg=Theme.BG_CARD,
                               highlightbackground=Theme.BORDER, highlightthickness=1)
        stats_card.pack(fill="x", padx=28, pady=(0, 12))

        self.alloc_stats_frame = tk.Frame(stats_card, bg=Theme.BG_CARD)
        self.alloc_stats_frame.pack(fill="x", padx=20, pady=14)

        self.alloc_stat_labels = {}
        stats = [
            ("pending",    "待配", Theme.WARNING),
            ("allocated", "已配", Theme.SUCCESS),
            ("assigned",  "指定", Theme.PRIMARY),
            ("no_stock",  "无库存", Theme.ERROR),
            ("boxes",     "可用箱", Theme.INFO),
        ]
        for key, label, color in stats:
            lbl = tk.Label(self.alloc_stats_frame, text=f"● {label}：0",
                           font=("Microsoft YaHei UI", 9, "bold"),
                           fg=color, bg=Theme.BG_CARD)
            lbl.pack(side="left", padx=(0, 20))
            self.alloc_stat_labels[key] = lbl

        # 工具栏
        toolbar = tk.Frame(page, bg=Theme.BG_PAGE)
        toolbar.pack(fill="x", padx=28, pady=(0, 8))

        ttk.Button(toolbar, text="📥 导入ERP订单", style="Secondary.TButton",
                    command=self._alloc_import_erp).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text="📋 导入ASN", style="Secondary.TButton",
                    command=self._alloc_import_asn).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text="📦 指定箱号", style="Secondary.TButton",
                    command=self._alloc_assign_box).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text="📋 指定DI", style="Secondary.TButton",
                    command=self._alloc_assign_di).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text="🗑️ 清除指定", style="Secondary.TButton",
                    command=self._alloc_clear_assign).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text="📤 导出配箱清单", style="Primary.TButton",
                    command=self._alloc_export).pack(side="left", padx=(20, 6))
        ttk.Button(toolbar, text="📧 一键发邮件", style="Primary.TButton",
                    command=self._alloc_send_email).pack(side="left", padx=(6, 0))

        # 表格区
        table_card = tk.Frame(page, bg=Theme.BG_CARD,
                              highlightbackground=Theme.BORDER, highlightthickness=1)
        table_card.pack(fill="both", expand=True, padx=28, pady=(0, 16))

        table_header = tk.Frame(table_card, bg=Theme.BG_CARD)
        table_header.pack(fill="x", padx=16, pady=(10, 4))
        tk.Label(table_header, text="配箱结果",
                 font=("Microsoft YaHei UI", 10, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        # Treeview 表格
        columns = ("order_no", "sales_org", "brand", "qty", "box_no", "di", "batch", "device", "status")
        self.alloc_tree = ttk.Treeview(
            table_card, columns=columns, show="headings", height=18,
            selectmode="extended"
        )
        # 列标题
        col_names = {
            "order_no":  "客户订单号",
            "sales_org": "销售组织",
            "brand":     "牌号",
            "qty":       "数量",
            "box_no":    "箱号",
            "di":        "DI",
            "batch":     "批次",
            "device":    "装置",
            "status":    "状态",
        }
        col_widths = {
            "order_no":  100,
            "sales_org": 80,
            "brand":     160,
            "qty":       50,
            "box_no":    100,
            "di":        120,
            "batch":      80,
            "device":     80,
            "status":     60,
        }
        for col in columns:
            self.alloc_tree.heading(col, text=col_names[col])
            self.alloc_tree.column(col, width=col_widths[col], minwidth=40, anchor="center")

        # 滚动条
        tree_scroll_y = ttk.Scrollbar(table_card, command=self.alloc_tree.yview)
        tree_scroll_x = ttk.Scrollbar(table_card, orient="horizontal", command=self.alloc_tree.xview)
        self.alloc_tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

        tree_scroll_y.pack(side="right", fill="y", padx=(0, 4), pady=(0, 8))
        tree_scroll_x.pack(side="bottom", fill="x", padx=(4, 4), pady=(0, 4))
        self.alloc_tree.pack(fill="both", expand=True, padx=(4, 0), pady=(0, 4))

        # 日志区
        log_card = tk.Frame(page, bg=Theme.BG_CARD,
                            highlightbackground=Theme.BORDER, highlightthickness=1)
        log_card.pack(fill="x", padx=28, pady=(0, 16))

        log_header = tk.Frame(log_card, bg=Theme.BG_CARD)
        log_header.pack(fill="x", padx=16, pady=(10, 4))
        tk.Label(log_header, text="操作记录",
                 font=("Microsoft YaHei UI", 10, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_CARD).pack(side="left")

        self.alloc_log_text = tk.Text(
            log_card, height=6, state="disabled",
            font=("Consolas", 9), wrap="word",
            bg="#FAFBFC", fg=Theme.TEXT_PRIMARY,
            borderwidth=0, highlightthickness=0,
            padx=12, pady=8
        )
        log_scroll = ttk.Scrollbar(log_card, command=self.alloc_log_text.yview)
        self.alloc_log_text.configure(yscrollcommand=log_scroll.set)
        log_scroll.pack(side="right", fill="y", padx=(0, 4), pady=(0, 8))
        self.alloc_log_text.pack(fill="x", padx=(12, 0), pady=(0, 8))

        self.alloc_log_text.tag_configure("red",    foreground=Theme.ERROR)
        self.alloc_log_text.tag_configure("green",  foreground=Theme.SUCCESS)
        self.alloc_log_text.tag_configure("warn",   foreground=Theme.WARNING)
        self.alloc_log_text.tag_configure("header", foreground=Theme.TEXT_SECONDARY)

        self.pages["allocation"] = page

    # ════════════════════════════════════════════════════
    #  配箱工具 - 数据加载
    # ════════════════════════════════════════════════════

    def _load_allocation_data(self):
        """加载配箱数据到表格"""
        file_path = self.alloc_file_var.get()
        if not file_path or not os.path.exists(file_path):
            return

        try:
            from allocation_module import load_allocation_workbook, get_current_orders, get_box_summary

            wb = load_allocation_workbook(file_path)
            self._alloc_orders = get_current_orders(wb)
            self._alloc_boxes  = get_box_summary(wb)
            wb.close()

            # 更新表格
            self._refresh_alloc_tree()

            self._alloc_log(f"✅ 数据加载成功：{len(self._alloc_orders)} 个订单，{len(self._alloc_boxes)} 个可用箱")

        except Exception as e:
            self._alloc_log(f"❌ 加载失败：{e}", "red")
            import traceback
            self._alloc_log(traceback.format_exc(), "red")

    def _refresh_alloc_tree(self):
        """刷新配箱表格"""
        # 清空
        for item in self.alloc_tree.get_children():
            self.alloc_tree.delete(item)

        if not hasattr(self, "_alloc_orders"):
            return

        # 填入数据
        for order in self._alloc_orders:
            status_color = {
                "待配": Theme.WARNING,
                "已配": Theme.SUCCESS,
                "指定": Theme.PRIMARY,
                "无库存": Theme.ERROR,
            }.get(order["status"], Theme.TEXT_PRIMARY)

            item_id = self.alloc_tree.insert("", "end", values=(
                order["order_no"],
                order["sales_org"],
                order["brand"],
                order["qty"],
                order["box_no"] or "",
                order["di"],
                order["batch"],
                order["device"],
                order["status"],
            ))
            # 状态列着色
            self.alloc_tree.set(item_id, "status", order["status"])
            # 用tag实现行着色（简化：只改状态文字颜色）

        # 更新统计
        self._update_alloc_stats()

    def _update_alloc_stats(self):
        if not hasattr(self, "_alloc_orders"):
            return
        orders = self._alloc_orders
        counts = {
            "pending":   sum(1 for o in orders if o["status"] == "待配"),
            "allocated": sum(1 for o in orders if o["status"] == "已配"),
            "assigned":  sum(1 for o in orders if o["status"] == "指定"),
            "no_stock":  sum(1 for o in orders if o["status"] == "无库存"),
            "boxes":     len(getattr(self, "_alloc_boxes", [])),
        }
        for key, lbl in self.alloc_stat_labels.items():
            lbl.config(text=f"● {lbl.cget('text').split('：')[0]}：{counts.get(key, 0)}")

    def _alloc_log(self, msg, tag=None):
        """写操作日志"""
        self.alloc_log_text.config(state="normal")
        if tag:
            self.alloc_log_text.insert("end", msg + "\n", tag)
        else:
            self.alloc_log_text.insert("end", msg + "\n")
        self.alloc_log_text.see("end")
        self.alloc_log_text.config(state="disabled")

    # ════════════════════════════════════════════════════
    #  配箱工具 - 操作功能
    # ════════════════════════════════════════════════════

    def _alloc_import_erp(self):
        """导入ERP订单"""
        if not self.alloc_file_var.get() or not os.path.exists(self.alloc_file_var.get()):
            messagebox.showwarning("提示", "请先选择配箱表文件")
            return

        erp_path = filedialog.askopenfilename(
            title="选择ERP导出的Excel文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls")])
        if not erp_path:
            return

        try:
            from allocation_module import import_erp_orders
            count, err = import_erp_orders(erp_path, self.alloc_file_var.get())
            if err:
                self._alloc_log(f"❌ 导入失败：{err}", "red")
                messagebox.showerror("导入失败", err)
            else:
                self._alloc_log(f"✅ 成功导入 {count} 个订单", "green")
                messagebox.showinfo("导入成功", f"成功导入 {count} 个订单")
                self._load_allocation_data()
        except Exception as e:
            self._alloc_log(f"❌ 导入异常：{e}", "red")

    def _alloc_import_asn(self):
        """导入ASN数据"""
        if not self.alloc_file_var.get() or not os.path.exists(self.alloc_file_var.get()):
            messagebox.showwarning("提示", "请先选择配箱表文件")
            return

        asn_path = filedialog.askopenfilename(
            title="选择ERP导出的ASN文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls")])
        if not asn_path:
            return

        try:
            from allocation_module import import_asn
            ok, err = import_asn(asn_path, self.alloc_file_var.get())
            if err:
                self._alloc_log(f"❌ ASN导入失败：{err}", "red")
            else:
                self._alloc_log(f"✅ ASN数据导入成功", "green")
                self._load_allocation_data()
        except Exception as e:
            self._alloc_log(f"❌ ASN导入异常：{e}", "red")

    def _alloc_assign_box(self):
        """指定箱号"""
        selected = self.alloc_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要指定箱号的订单行")
            return

        if not hasattr(self, "_alloc_boxes") or not self._alloc_boxes:
            messagebox.showwarning("提示", "没有可用箱子，请先加载数据")
            return

        # 获取选中订单的信息
        selected_orders = []
        for item_id in selected:
            vals = self.alloc_tree.item(item_id, "values")
            selected_orders.append({
                "item_id": item_id,
                "order_no": vals[0],
                "sales_org": vals[1],
                "brand": vals[2],
                "qty": vals[3],
            })

        # 构建匹配的箱号列表
        # 优先显示同牌号+同Hub的箱子
        if selected_orders:
            match_brand = selected_orders[0]["brand"]
            match_hub = selected_orders[0]["sales_org"]
        else:
            match_brand = ""
            match_hub = ""

        matched_boxes = []
        other_boxes = []
        for b in self._alloc_boxes:
            b_info = f"{b['box_no']}  |  {b['brand'][:20]}  |  {b['hub']}  |  ETA={b.get('eta','')}"
            if b["brand"] == match_brand and b["hub"] == match_hub:
                matched_boxes.append((b, b_info))
            else:
                other_boxes.append((b, b_info))

        # 弹窗选择箱号
        dlg = tk.Toplevel(self.root)
        dlg.title("指定箱号")
        dlg.geometry("700x500")
        dlg.transient(self.root)
        dlg.grab_set()

        # 已选订单信息
        info_frame = tk.Frame(dlg, bg=Theme.BG_CARD)
        info_frame.pack(fill="x", padx=16, pady=12)
        tk.Label(info_frame, text=f"已选 {len(selected_orders)} 个订单",
                 font=("Microsoft YaHei UI", 10, "bold"),
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(anchor="w", padx=12, pady=(8,4))
        for o in selected_orders[:5]:
            tk.Label(info_frame, text=f"  {o['order_no']} | {o['sales_org']} | {o['brand'][:25]} | {o['qty']}吨",
                     font=("Microsoft YaHei UI", 9), bg=Theme.BG_CARD, fg=Theme.TEXT_SECONDARY).pack(anchor="w", padx=12)
        if len(selected_orders) > 5:
            tk.Label(info_frame, text=f"  ... 等{len(selected_orders)}个",
                     font=("Microsoft YaHei UI", 9), bg=Theme.BG_CARD, fg=Theme.TEXT_SECONDARY).pack(anchor="w", padx=12)

        # 匹配箱子列表
        tk.Label(dlg, text=f"匹配的箱子（{len(matched_boxes)}个）",
                 font=("Microsoft YaHei UI", 9, "bold"),
                 bg=Theme.BG_PAGE, fg=Theme.SUCCESS).pack(anchor="w", padx=20, pady=(8,2))

        listbox = tk.Listbox(dlg, font=("Consolas", 9), height=12,
                             selectmode="single", bg="#FAFBFC")
        for b, info in matched_boxes:
            listbox.insert("end", info)
        listbox.pack(fill="both", expand=True, padx=20, pady=(0,8))

        # 其他箱子
        tk.Label(dlg, text=f"其他箱子（{len(other_boxes)}个）",
                 font=("Microsoft YaHei UI", 9),
                 bg=Theme.BG_PAGE, fg=Theme.TEXT_SECONDARY).pack(anchor="w", padx=20, pady=(0,2))

        other_listbox = tk.Listbox(dlg, font=("Consolas", 9), height=5,
                                   selectmode="single", bg="#FAFBFC")
        for b, info in other_boxes:
            other_listbox.insert("end", info)
        other_listbox.pack(fill="x", padx=20, pady=(0,8))

        # 按钮
        btn_frame = tk.Frame(dlg, bg=Theme.BG_PAGE)
        btn_frame.pack(fill="x", padx=20, pady=(0,16))

        def _confirm():
            sel_idx = listbox.curselection()
            other_idx = other_listbox.curselection()

            if sel_idx:
                box = matched_boxes[sel_idx[0]][0]
            elif other_idx:
                box = other_boxes[other_idx[0]][0]
            else:
                messagebox.showwarning("提示", "请选择一个箱子")
                return

            # 写入J列
            self._do_assign_box(selected_orders, box)
            dlg.destroy()

        ttk.Button(btn_frame, text="确认指定", style="Primary.TButton",
                    command=_confirm).pack(side="left")
        ttk.Button(btn_frame, text="取消", style="Secondary.TButton",
                    command=dlg.destroy).pack(side="left", padx=(8,0))

    def _do_assign_box(self, selected_orders, box):
        """执行指定箱号写入"""
        try:
            from allocation_module import load_allocation_workbook
            from openpyxl import load_workbook as load_wb_rw

            alloc_path = self.alloc_file_var.get()
            # 用公式版本打开（可写入）
            wb = load_wb_rw(alloc_path)
            ws = wb["配箱公式"]

            # 找到选中订单的行号并写入J列
            for o in selected_orders:
                order_no = o["order_no"]
                for row in range(2, ws.max_row + 1):
                    z_val = ws.cell(row=row, column=26).value
                    if z_val and str(int(z_val) if isinstance(z_val, (int,float)) else z_val) == order_no:
                        ws.cell(row=row, column=10).value = box["box_no"]
                        break

            wb.save(alloc_path)
            wb.close()

            self._alloc_log(f"✅ 已指定箱号 {box['box_no']}", "green")
            self._load_allocation_data()

        except Exception as e:
            self._alloc_log(f"❌ 指定箱号失败：{e}", "red")
            messagebox.showerror("错误", str(e))

    def _alloc_assign_di(self):
        """指定DI"""
        selected = self.alloc_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要指定DI的订单行")
            return

        if not hasattr(self, "_alloc_boxes") or not self._alloc_boxes:
            messagebox.showwarning("提示", "没有可用箱子，请先加载数据")
            return

        # 收集所有可用DI
        di_map = {}
        for b in self._alloc_boxes:
            if b["di"] and b["di"] not in di_map:
                di_map[b["di"]] = {
                    "di": b["di"],
                    "hub": b["hub"],
                    "brand": b["brand"],
                    "count": 0,
                    "eta": b.get("eta"),
                }
            if b["di"]:
                di_map[b["di"]]["count"] += 1

        di_list = sorted(di_map.values(), key=lambda x: x["eta"] or datetime(2099,1,1))

        # 弹窗选择DI
        dlg = tk.Toplevel(self.root)
        dlg.title("指定DI")
        dlg.geometry("600x400")
        dlg.transient(self.root)
        dlg.grab_set()

        selected_orders = []
        for item_id in selected:
            vals = self.alloc_tree.item(item_id, "values")
            selected_orders.append(vals)

        tk.Label(dlg, text=f"已选 {len(selected_orders)} 个订单，请选择DI：",
                 font=("Microsoft YaHei UI", 10, "bold"),
                 bg=Theme.BG_PAGE, fg=Theme.TEXT_PRIMARY).pack(anchor="w", padx=20, pady=(16,8))

        listbox = tk.Listbox(dlg, font=("Consolas", 9), height=12,
                             selectmode="single", bg="#FAFBFC")
        for di_info in di_list:
            eta_str = di_info["eta"].strftime("%Y/%m/%d") if isinstance(di_info["eta"], datetime) else ""
            line = f"{di_info['di']}  |  {di_info['hub']}  |  {di_info['brand'][:20]}  |  {di_info['count']}箱  |  ETA={eta_str}"
            listbox.insert("end", line)
        listbox.pack(fill="both", expand=True, padx=20, pady=(0,8))

        btn_frame = tk.Frame(dlg, bg=Theme.BG_PAGE)
        btn_frame.pack(fill="x", padx=20, pady=(0,16))

        def _confirm():
            sel_idx = listbox.curselection()
            if not sel_idx:
                messagebox.showwarning("提示", "请选择一个DI")
                return

            di_info = di_list[sel_idx[0]]
            di = di_info["di"]

            # 找该DI下的箱子，按ETA排序
            di_boxes = [b for b in self._alloc_boxes if b["di"] == di]
            di_boxes.sort(key=lambda x: x.get("eta") or datetime(2099,1,1))

            if not di_boxes:
                messagebox.showwarning("提示", f"DI {di} 下没有可用箱子")
                return

            # 按序分配
            self._do_assign_di(selected_orders, di_boxes)
            dlg.destroy()

        ttk.Button(btn_frame, text="确认指定", style="Primary.TButton",
                    command=_confirm).pack(side="left")
        ttk.Button(btn_frame, text="取消", style="Secondary.TButton",
                    command=dlg.destroy).pack(side="left", padx=(8,0))

    def _do_assign_di(self, selected_orders, di_boxes):
        """执行指定DI写入"""
        try:
            from openpyxl import load_workbook as load_wb_rw

            alloc_path = self.alloc_file_var.get()
            wb = load_wb_rw(alloc_path)
            ws = wb["配箱公式"]

            box_idx = 0
            for item_vals in selected_orders:
                order_no = item_vals[0]
                for row in range(2, ws.max_row + 1):
                    z_val = ws.cell(row=row, column=26).value
                    if z_val and str(int(z_val) if isinstance(z_val, (int,float)) else z_val) == order_no:
                        if box_idx < len(di_boxes):
                            ws.cell(row=row, column=10).value = di_boxes[box_idx]["box_no"]
                            box_idx += 1
                        break

            wb.save(alloc_path)
            wb.close()

            self._alloc_log(f"✅ 已指定DI，分配了 {box_idx} 个箱子", "green")
            self._load_allocation_data()

        except Exception as e:
            self._alloc_log(f"❌ 指定DI失败：{e}", "red")

    def _alloc_clear_assign(self):
        """清除指定（恢复J列公式）"""
        selected = self.alloc_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要清除指定的订单行")
            return

        if not messagebox.askyesno("确认", "确定要清除选中订单的指定箱号吗？"):
            return

        try:
            from openpyxl import load_workbook as load_wb_rw

            alloc_path = self.alloc_file_var.get()
            wb = load_wb_rw(alloc_path)
            ws = wb["配箱公式"]

            cleared = 0
            for item_id in selected:
                vals = self.alloc_tree.item(item_id, "values")
                order_no = vals[0]
                for row in range(2, ws.max_row + 1):
                    z_val = ws.cell(row=row, column=26).value
                    if z_val and str(int(z_val) if isinstance(z_val, (int,float)) else z_val) == order_no:
                        # 清空J列（设为None，让公式重新计算）
                        ws.cell(row=row, column=10).value = None
                        cleared += 1
                        break

            wb.save(alloc_path)
            wb.close()

            self._alloc_log(f"✅ 已清除 {cleared} 个订单的指定箱号", "green")
            self._load_allocation_data()

        except Exception as e:
            self._alloc_log(f"❌ 清除指定失败：{e}", "red")

    def _alloc_export(self):
        """导出配箱清单"""
        if not hasattr(self, "_alloc_orders"):
            messagebox.showwarning("提示", "请先加载数据")
            return

        output_path = filedialog.asksaveasfilename(
            title="导出配箱清单",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=f"配箱清单_{datetime.now().strftime('%Y%m%d')}.xlsx")
        if not output_path:
            return

        try:
            from allocation_module import export_allocation_list
            export_allocation_list(self._alloc_orders, output_path)
            self._alloc_log(f"✅ 配箱清单已导出：{os.path.basename(output_path)}", "green")
            messagebox.showinfo("导出成功", f"配箱清单已保存到：\n{output_path}")
        except Exception as e:
            self._alloc_log(f"❌ 导出失败：{e}", "red")

    def _alloc_send_email(self):
        """一键发邮件（改单邮件+发货通知）"""
        self._alloc_log("📧 邮件功能：请切换到「改单邮件」和「发货通知」页面发送")
        messagebox.showinfo("提示",
            "邮件发送功能请使用左侧导航的：\n\n"
            "📋 改单邮件 — 批量发送改单邮件+COA附件\n"
            "🚚 发货通知 — 发送发货通知邮件\n\n"
            "配箱结果会自动同步到邮件数据。")

    # ───── 文件选择回调 ─────

    def _select_order_excel(self):
        p = filedialog.askopenfilename(title="选择 Excel 文件",
                                        filetypes=[("Excel 文件", "*.xlsx *.xls")])
        if p:
            self.order_excel_path.set(p)
            if not self.order_output_dir.get():
                self.order_output_dir.set(os.path.join(os.path.dirname(p), "output"))
            self._save_current_config()

    def _select_order_coa(self):
        p = filedialog.askdirectory(title="选择 COA 文件夹")
        if p:
            self.order_coa_dir.set(p)
            self._save_current_config()

    def _select_order_pdf(self):
        p = filedialog.askopenfilename(title="选择 PDF 文件",
                                        filetypes=[("PDF 文件", "*.pdf")])
        if p:
            self.order_pdf_path.set(p)
            self._save_current_config()

    def _select_order_output(self):
        p = filedialog.askdirectory(title="选择输出目录")
        if p:
            self.order_output_dir.set(p)
            self._save_current_config()

    def _select_ship_dir(self):
        p = filedialog.askdirectory(title="选择搜索文件夹")
        if p:
            self.ship_search_dir.set(p)
            self._save_current_config()

    # ───── 日志辅助 ─────

    def _log(self, text_widget, msg, color=None):
        text_widget.config(state="normal")
        if color:
            text_widget.insert("end", msg + "\n", color)
        else:
            text_widget.insert("end", msg + "\n")
        text_widget.see("end")
        text_widget.config(state="disabled")
        text_widget.update_idletasks()

    # ════════════════════════════════════════════════════════
    #  确认预览对话框
    # ════════════════════════════════════════════════════════

    def _show_confirm_dialog(self, title, preview_lines):
        """显示发送前确认对话框，返回 True/False"""
        dialog = tk.Toplevel(self.root)
        dialog.title("发送确认")
        dialog.geometry("520x420")
        dialog.resizable(False, False)
        dialog.configure(bg=Theme.BG_PAGE)
        dialog.transient(self.root)
        dialog.grab_set()

        # 居中显示
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 520) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 420) // 2
        dialog.geometry(f"+{x}+{y}")

        # 标题
        tk.Label(dialog, text=title,
                 font=("Microsoft YaHei UI", 13, "bold"),
                 fg=Theme.TEXT_PRIMARY, bg=Theme.BG_PAGE).pack(padx=24, pady=(20, 4), anchor="w")

        tk.Label(dialog, text="请确认以下内容无误后发送：",
                 font=("Microsoft YaHei UI", 9),
                 fg=Theme.TEXT_SECONDARY, bg=Theme.BG_PAGE).pack(padx=24, pady=(0, 12), anchor="w")

        # 预览区
        preview_frame = tk.Frame(dialog, bg=Theme.BG_CARD,
                                 highlightbackground=Theme.BORDER,
                                 highlightthickness=1)
        preview_frame.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        preview_text = tk.Text(preview_frame, height=12, state="normal",
                               font=("Consolas", 9), wrap="word",
                               bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
                               borderwidth=0, highlightthickness=0,
                               padx=12, pady=8)
        scrollbar = ttk.Scrollbar(preview_frame, command=preview_text.yview)
        preview_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y", padx=(0, 4), pady=4)
        preview_text.pack(fill="both", expand=True, padx=4, pady=4)

        preview_text.tag_configure("green", foreground=Theme.SUCCESS)
        preview_text.tag_configure("red",   foreground=Theme.ERROR)
        preview_text.tag_configure("warn",  foreground=Theme.WARNING)
        preview_text.tag_configure("bold",  font=("Consolas", 9, "bold"))

        for line, color in preview_lines:
            if color:
                preview_text.insert("end", line + "\n", color)
            else:
                preview_text.insert("end", line + "\n")
        preview_text.config(state="disabled")

        # 按钮区
        btn_frame = tk.Frame(dialog, bg=Theme.BG_PAGE)
        btn_frame.pack(fill="x", padx=24, pady=(0, 20))

        result = [False]

        def on_cancel():
            result[0] = False
            dialog.destroy()

        def on_confirm():
            result[0] = True
            dialog.destroy()

        ttk.Button(btn_frame, text="取消", style="Secondary.TButton",
                   command=on_cancel).pack(side="right", padx=(8, 0))
        ttk.Button(btn_frame, text="✓  确认发送", style="Primary.TButton",
                   command=on_confirm).pack(side="right")

        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        dialog.wait_window()

        return result[0]

    # ════════════════════════════════════════════════════════
    #  改单邮件执行
    # ════════════════════════════════════════════════════════

    def _on_order_execute(self):
        if self._running:
            return

        # 检查必填项
        missing = []
        if not self.order_excel_path.get() or not os.path.isfile(self.order_excel_path.get()):
            missing.append("Excel 文件")
        if not self.order_coa_dir.get() or not os.path.isdir(self.order_coa_dir.get()):
            missing.append("COA 文件夹")
        if not self.order_pdf_path.get() or not os.path.isfile(self.order_pdf_path.get()):
            missing.append("PDF 文件")
        if not self.order_output_dir.get():
            missing.append("输出目录")
        if not self.order_email_user.get():
            missing.append("邮箱账号")
        if not self.order_email_pass.get():
            missing.append("授权码")
        if not self.order_recipients.get():
            missing.append("收件人")

        if missing:
            messagebox.showerror("缺少必填项",
                                 "请先在设置中配置：\n" + "\n".join(f"  • {m}" for m in missing))
            return

        self._save_current_config()

        # 发送前确认
        if self.confirm_before_send.get():
            try:
                email_data, order_batch = read_email_data(self.order_excel_path.get())
                pdf_map = split_pdf_by_order(self.order_pdf_path.get(), self.order_output_dir.get())
            except Exception as e:
                messagebox.showerror("预览失败", f"读取数据时出错：{e}")
                return

            preview_lines = []
            preview_lines.append(("收件人：" + self.order_recipients.get(), None))
            preview_lines.append(("邮箱账号：" + self.order_email_user.get(), None))
            preview_lines.append(("", None))
            preview_lines.append(("── 待发送邮件 ──", "bold"))

            valid_count = 0
            for item in email_data:
                order_no = item["订单号"]
                pdf_info = pdf_map.get(order_no)
                if pdf_info:
                    page_count = pdf_info["pages"]
                    coa_files = find_coa_files(order_no, self.order_coa_dir.get(), order_batch)
                    coa_status = "✅" if coa_files else "⚠️ COA缺失"
                    preview_lines.append((f"PO {order_no}（{page_count}页）{coa_status}", "green" if coa_files else "warn"))
                    valid_count += 1
                else:
                    preview_lines.append((f"PO {order_no} ❌ PDF中无此订单", "red"))

            preview_lines.append(("", None))
            preview_lines.append((f"共 {valid_count} 封邮件将被发送", "bold"))

            if not self._show_confirm_dialog("改单邮件 · 发送确认", preview_lines):
                return

        self._running = True
        self.order_btn_run.config(state="disabled")
        self.order_progress["value"] = 0
        self.order_progress_label.config(text="")
        threading.Thread(target=self._do_order_task, daemon=True).start()

    def _do_order_task(self):
        log = self.order_log_text

        # 清空日志
        log.config(state="normal")
        log.delete("1.0", "end")
        log.config(state="disabled")

        try:
            excel_path   = self.order_excel_path.get()
            coa_dir      = self.order_coa_dir.get()
            pdf_path     = self.order_pdf_path.get()
            output_dir   = self.order_output_dir.get()
            smtp_server  = self.order_smtp_server.get()
            smtp_port    = self.order_smtp_port.get()
            username     = self.order_email_user.get()
            password     = self.order_email_pass.get()
            recipients   = self.order_recipients.get()

            os.makedirs(output_dir, exist_ok=True)

            # Step 1：读取 Excel
            email_data, order_batch = read_email_data(excel_path)
            total = len(email_data)

            # Step 2：拆分 PDF
            self._log(log, f"正在拆分 PDF，共 {total} 个PO待处理…")
            pdf_map = split_pdf_by_order(pdf_path, output_dir)

            # Step 3：逐订单处理（复用同一 SMTP 连接）
            self.order_progress["maximum"] = total
            smtp_conn = None  # 复用的 SMTP 连接

            for i, item in enumerate(email_data, 1):
                order_no = item["订单号"]

                # 更新进度
                self.order_progress["value"] = i
                self.order_progress_label.config(text=f"{i}/{total}")

                # 检查 PDF
                pdf_info = pdf_map.get(order_no)
                if not pdf_info or not os.path.exists(pdf_info["path"]):
                    self._log(log, f"PO {order_no} ❌ PDF中无此订单，已跳过", "red")
                    continue

                page_count = pdf_info["pages"]

                # COA 打包
                coa_files = find_coa_files(order_no, coa_dir, order_batch)
                zip_path = os.path.join(output_dir, f"{order_no}.zip")
                coa_ok = bool(coa_files)

                if coa_ok:
                    zip_files(coa_files, zip_path)

                # 邮件标题
                subject = item["邮件标题"]
                if not coa_ok:
                    subject = f"【COA待确认】{subject}"

                # 邮件正文
                body = item["邮件正文"]

                # 附件
                attachments = [pdf_info["path"]]
                if coa_ok and os.path.exists(zip_path):
                    attachments.append(zip_path)

                # 发送（复用连接，失败时自动重连重试）
                max_send_retries = 2
                for send_attempt in range(1, max_send_retries + 1):
                    try:
                        smtp_conn = send_email(
                            smtp_server, smtp_port, username, password,
                            recipients, subject, body, attachments,
                            server=smtp_conn)
                        coa_tag = "" if coa_ok else " ⚠️COA缺失"
                        self._log(log, f"PO {order_no}（{page_count}页）✅{coa_tag}", "green")
                        break  # 发送成功，跳出重试循环
                    except Exception as e:
                        # 连接已废弃，下次重试会新建连接
                        smtp_conn = None
                        if send_attempt < max_send_retries:
                            self._log(log, f"PO {order_no}（{page_count}页）⚠️ 发送失败，正在重连重试…", "warn")
                            import time
                            time.sleep(2)
                        else:
                            self._log(log, f"PO {order_no}（{page_count}页）❌ 发送失败：{e}", "red")

            # 关闭 SMTP 连接
            if smtp_conn:
                try:
                    smtp_conn.quit()
                except Exception:
                    pass

            self.order_progress_label.config(text=f"{total}/{total}")

        except Exception as e:
            self._log(log, f"❌ 执行出错：{e}", "red")
            import traceback
            self._log(log, traceback.format_exc())

        finally:
            self._running = False
            self.order_btn_run.config(state="normal")

    # ════════════════════════════════════════════════════════
    #  发货通知执行
    # ════════════════════════════════════════════════════════

    def _on_ship_execute(self):
        if self._running:
            return

        # 检查必填项
        missing = []
        if not self.ship_search_dir.get() or not os.path.isdir(self.ship_search_dir.get()):
            missing.append("搜索文件夹")
        if not self.ship_keyword.get().strip():
            missing.append("文件特征值")
        if not self.ship_subject_pre.get().strip():
            missing.append("邮件标题前缀")
        if not self.ship_email_user.get():
            missing.append("邮箱账号")
        if not self.ship_email_pass.get():
            missing.append("授权码")
        if not self.ship_recipients.get():
            missing.append("收件人")

        if missing:
            messagebox.showerror("缺少必填项",
                                 "请先在设置中配置：\n" + "\n".join(f"  • {m}" for m in missing))
            return

        self._save_current_config()

        # 发送前确认
        if self.confirm_before_send.get():
            keyword = self.ship_keyword.get().strip()
            search_dir = self.ship_search_dir.get()
            found_file = find_newest_file_by_keyword(search_dir, keyword)

            preview_lines = []
            preview_lines.append(("收件人：" + self.ship_recipients.get(), None))
            preview_lines.append(("邮箱账号：" + self.ship_email_user.get(), None))
            preview_lines.append(("", None))
            preview_lines.append(("── 发货通知预览 ──", "bold"))

            now = datetime.now()
            dt_str = now.strftime("%Y%m%d-%H:%M")
            subject = f"{self.ship_subject_pre.get()} {dt_str}"
            preview_lines.append((f"邮件标题：{subject}", None))

            if found_file:
                fname = os.path.basename(found_file)
                mtime = datetime.fromtimestamp(os.path.getmtime(found_file))
                preview_lines.append((f"附件文件：{fname}", "green"))
                preview_lines.append((f"文件修改时间：{mtime.strftime('%Y-%m-%d %H:%M')}", None))
            else:
                preview_lines.append((f"⚠️ 在文件夹中未找到包含「{keyword}」的文件", "red"))

            if not self._show_confirm_dialog("发货通知 · 发送确认", preview_lines):
                return

        self._running = True
        self.ship_btn_run.config(state="disabled")
        self.ship_progress["value"] = 0
        self.ship_progress_label.config(text="")
        threading.Thread(target=self._do_ship_task, daemon=True).start()

    def _do_ship_task(self):
        log = self.ship_log_text

        # 清空日志
        log.config(state="normal")
        log.delete("1.0", "end")
        log.config(state="disabled")

        try:
            keyword     = self.ship_keyword.get().strip()
            search_dir  = self.ship_search_dir.get()
            subject_pre = self.ship_subject_pre.get().strip()
            body        = self.ship_email_body.get() or "您好，请查收附件。"
            smtp_server = self.ship_smtp_server.get()
            smtp_port   = self.ship_smtp_port.get()
            username    = self.ship_email_user.get()
            password    = self.ship_email_pass.get()
            recipients  = self.ship_recipients.get()

            # 查找文件
            self._log(log, f"正在搜索包含「{keyword}」的文件…")
            found_file = find_newest_file_by_keyword(search_dir, keyword)

            if not found_file:
                self._log(log, f"❌ 未找到包含「{keyword}」的文件", "red")
                return

            fname = os.path.basename(found_file)
            self._log(log, f"找到文件：{fname}")

            # 构建邮件
            now = datetime.now()
            dt_str = now.strftime("%Y%m%d-%H:%M")
            subject = f"{subject_pre} {dt_str}"

            self._log(log, f"邮件标题：{subject}")
            self._log(log, f"正在发送…")

            self.ship_progress["value"] = 50
            self.ship_progress_label.config(text="发送中…")

            smtp_conn = send_email(smtp_server, smtp_port, username, password,
                                   recipients, subject, body, [found_file])
            try:
                smtp_conn.quit()
            except Exception:
                pass

            self.ship_progress["value"] = 100
            self.ship_progress_label.config(text="完成")
            self._log(log, f"✅ 发送成功", "green")

        except Exception as e:
            self._log(log, f"❌ 发送失败：{e}", "red")
            import traceback
            self._log(log, traceback.format_exc())

        finally:
            self._running = False
            self.ship_btn_run.config(state="normal")


# ══════════════════════════════════════════════════════════
#  启动
# ══════════════════════════════════════════════════════════

if __name__ == "__main__":
    root = tk.Tk()
    app = EmailToolApp(root)
    root.mainloop()
