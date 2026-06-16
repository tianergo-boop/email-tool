"""
配箱工具核心逻辑模块
处理配箱表的读写、配箱计算、ERP/ASN数据导入导出
支持 xlwings（快速COM读取）+ openpyxl（纯Python回退）
"""

import os
import json
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# ══════════════════════════════════════════════════════
#  xlwings 可用性检测
# ══════════════════════════════════════════════════════

_xlwings_available = None

def has_xlwings():
    """检测 xlwings 是否可用（Excel/WPS 已安装且可调用）"""
    global _xlwings_available
    if _xlwings_available is not None:
        return _xlwings_available
    try:
        import xlwings as xw
        # 尝试创建一个不可见的应用实例来验证
        app = xw.App(visible=False)
        app.quit()
        _xlwings_available = True
    except Exception:
        _xlwings_available = False
    return _xlwings_available


# ══════════════════════════════════════════════════════
#  常量定义
# ══════════════════════════════════════════════════════

# 配箱表的 sheet 名
SHEET_FORMULA  = "配箱公式"   # ERP订单数据 + 配箱公式
SHEET_BOX      = "配箱表"     # 物流+ASN视图，真正配箱依据
SHEET_SUMMARY  = "汇总"       # 配箱表sheet的透视表
SHEET_ASN      = "ASN"        # ERP导出的ASN数据
SHEET_LOGISTICS = "物流跟踪"   # RPA每日更新的物流数据

# 「配箱公式」sheet 的列
COL_ERP_MATERIAL = "A"   # erp物料
COL_BRAND_MAP    = "H"   # 牌号（公式：VLOOKUP A:B）
COL_BOX_NO       = "J"   # 箱号（核心：VLOOKUP汇总表）
COL_PO           = "K"   # PO
COL_DI           = "S"   # DI（交货指示号）
# Z列开始：ERP原始数据复制区
COL_ORDER_NO     = "Z"   # 客户订单号
COL_STATUS       = "AA"  # 单据状态
COL_LINE_NO      = "AB"  # 行号
COL_SALES_ORG   = "AC"  # 销售组织（如"上海Hub"）
COL_MATERIAL    = "AD"  # 物料名称
COL_DEVICE       = "M"   # 装置
COL_COA          = "N"   # COA
COL_BATCH        = "BJ"  # 批次号
COL_QTY          = "AI"  # 数量
COL_SHIP_POINT   = "AT"  # 交货指示号（AT列）
COL_INCOTERM     = "X"   # SO Incoterm

# 「汇总」sheet 区域
SUMMARY_BRAND_COL  = "H"   # 牌号
SUMMARY_HUB_COL   = "B"   # Hub
SUMMARY_ETA_COL   = "J"   # ETA
SUMMARY_ETD_COL   = "K"   # ETD
SUMMARY_DI_COL     = "M"   # DI
SUMMARY_BOX_COL   = "O"   # 箱号
SUMMARY_SONO_COL  = "D"   # SONO

# 数据起始行（配箱公式sheet，第16行是表头）
DATA_START_ROW = 17

# 列号常量（1-indexed，用于 xlwings / openpyxl cell 访问）
COL_IDX = {
    "A": 1, "H": 8, "J": 10, "K": 11, "M": 13, "N": 14, "S": 19, "X": 24,
    "Z": 26, "AA": 27, "AB": 28, "AC": 29, "AD": 30, "AI": 35, "AT": 46, "BJ": 62,
}


# ══════════════════════════════════════════════════════
#  列映射配置
# ══════════════════════════════════════════════════════

DEFAULT_ERP_MAPPING = {
    "客户订单号": "Z",
    "单据状态": "AA",
    "行号": "AB",
    "销售组织": "AC",
    "物料名称": "AD",
    "LYB陆运承运商": "AE",
    "LYB承运商联系人": "AF",
    "规格型号": "AG",
    "计量单位": "AH",
    "数量": "AI",
    "单价": "AJ",
    "金额": "AK",
    "含税单价": "AL",
    "税额": "AM",
    "价税合计": "AN",
    "发货组织": "AO",
    "客户订单分录序号": "AQ",
    "牌号": "AR",
    "装船要求": "AS",
    "贸易条款": "X",
    "制单日期": "Y",
}


def get_mapping_config_path():
    """获取映射配置文件路径"""
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "allocation_mapping.json")


def load_mapping_config():
    """加载列映射配置"""
    path = get_mapping_config_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"erp_mapping": DEFAULT_ERP_MAPPING}


def save_mapping_config(cfg):
    path = get_mapping_config_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ══════════════════════════════════════════════════════
#  openpyxl 缓存
# ══════════════════════════════════════════════════════

_cache = {}

def _get_cache_key(file_path):
    """用文件路径 + mtime 作为缓存 key"""
    mtime = os.path.getmtime(file_path) if os.path.exists(file_path) else 0
    return f"{file_path}|{mtime}"

def _invalidate_cache(file_path):
    """文件写入后使缓存失效"""
    to_del = [k for k in _cache if k.startswith(file_path + "|")]
    for k in to_del:
        del _cache[k]


# ══════════════════════════════════════════════════════
#  加载配箱表
# ══════════════════════════════════════════════════════

def load_allocation_workbook(file_path):
    """
    加载配箱表
    优先使用 xlwings（COM，可读取公式计算结果），
    回退到 openpyxl data_only=True
    返回一个 WorkbookProxy 对象，统一 xlwings / openpyxl 接口
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在：{file_path}")

    if has_xlwings():
        try:
            return _XlwingsWorkbookProxy(file_path)
        except Exception:
            pass  # fallback

    return _OpenpyxlWorkbookProxy(file_path)


class _XlwingsWorkbookProxy:
    """xlwings 包装，提供与 openpyxl 兼容的读取接口"""

    def __init__(self, file_path):
        import xlwings as xw
        self._app = xw.App(visible=False)
        self._book = self._app.books.open(file_path)
        self._file_path = file_path
        self._sheets = {}
        # 预加载所有 sheet 数据
        for s in self._book.sheets:
            self._sheets[s.name] = _XlwingsSheetProxy(s)

    def __getitem__(self, name):
        if name not in self._sheets:
            raise KeyError(f"Sheet '{name}' 不存在")
        return self._sheets[name]

    def __contains__(self, name):
        return name in self._sheets

    def close(self):
        try:
            self._book.close()
        except Exception:
            pass
        try:
            self._app.quit()
        except Exception:
            pass

    @property
    def sheetnames(self):
        return list(self._sheets.keys())


class _XlwingsSheetProxy:
    """xlwings sheet 包装，模拟 openpyxl worksheet 的 cell 访问"""

    def __init__(self, xlwings_sheet):
        self._sheet = xlwings_sheet
        # 一次性读取全部数据到二维列表（行优先）
        used = xlwings_sheet.used_range
        if used and used.value:
            self._data = used.value
            self._row_offset = used.row - 1  # xlwings row 是 1-indexed
            self._col_offset = used.column - 1
        else:
            self._data = [[]]
            self._row_offset = 0
            self._col_offset = 0

    @property
    def max_row(self):
        return self._row_offset + len(self._data) if self._data else 1

    @property
    def max_column(self):
        if not self._data or not self._data[0]:
            return 1
        return self._col_offset + max(len(r) for r in self._data)

    def cell(self, row, column, value=None):
        """模拟 openpyxl 的 cell() 方法（1-indexed）"""
        if value is not None:
            # 写入模式：通过 xlwings 直接写
            self._sheet.range((row, column)).value = value
            # 更新缓存
            r = row - self._row_offset - 1
            c = column - self._col_offset - 1
            if 0 <= r < len(self._data) and 0 <= c < len(self._data[r]) if r < len(self._data) else False:
                self._data[r][c] = value
            return

        r = row - self._row_offset - 1
        c = column - self._col_offset - 1
        try:
            if 0 <= r < len(self._data):
                row_data = self._data[r]
                if isinstance(row_data, (list, tuple)) and 0 <= c < len(row_data):
                    return _CellValue(row_data[c])
            return _CellValue(None)
        except (IndexError, TypeError):
            return _CellValue(None)


class _CellValue:
    """模拟 openpyxl Cell 的 .value 属性"""
    def __init__(self, val):
        self.value = val
    def __repr__(self):
        return f"CellValue({self.value!r})"


class _OpenpyxlWorkbookProxy:
    """openpyxl 包装，统一接口"""

    def __init__(self, file_path):
        self._wb = load_workbook(file_path, data_only=True)
        self._file_path = file_path

    def __getitem__(self, name):
        return self._wb[name]

    def __contains__(self, name):
        return name in self._wb.sheetnames

    def close(self):
        self._wb.close()

    @property
    def sheetnames(self):
        return self._wb.sheetnames


# ══════════════════════════════════════════════════════
#  读取汇总数据
# ══════════════════════════════════════════════════════

def get_box_summary(wb_or_path):
    """
    读取「汇总」sheet，返回可用库存池
    支持 wb 对象或文件路径
    """
    if isinstance(wb_or_path, str):
        wb = load_allocation_workbook(wb_or_path)
        try:
            return get_box_summary(wb)
        finally:
            wb.close()

    # 检查缓存
    if isinstance(wb_or_path, _OpenpyxlWorkbookProxy):
        cache_key = _get_cache_key(wb_or_path._file_path) + "|summary"
        if cache_key in _cache:
            return _cache[cache_key]

    ws = wb_or_path["汇总"]
    boxes = []

    seen_boxes = set()
    for row in range(6, ws.max_row + 1):
        box_no = ws.cell(row=row, column=15).value  # O列 = 箱号
        if not box_no or str(box_no).strip() == "":
            continue

        box_str = str(box_no).strip()
        if box_str in seen_boxes:
            continue
        seen_boxes.add(box_str)

        hub    = ws.cell(row=row, column=2).value   # B列 = Hub
        brand  = ws.cell(row=row, column=8).value   # H列 = 产品描述/牌号
        di     = ws.cell(row=row, column=13).value  # M列 = DI
        eta    = ws.cell(row=row, column=10).value   # J列 = ETA
        etd    = ws.cell(row=row, column=11).value   # K列 = ETD
        atd    = ws.cell(row=row, column=9).value    # I列 = ATD
        weight = ws.cell(row=row, column=16).value   # P列 = 重量
        batch  = ws.cell(row=row, column=17).value   # Q列 = 批次
        device = ws.cell(row=row, column=19).value   # S列 = 装置
        waybill= ws.cell(row=row, column=3).value    # C列 = 运单号

        key = f"{weight}|{hub}|{brand}"

        boxes.append({
            "box_no": box_str,
            "brand": str(brand) if brand else "",
            "hub": str(hub) if hub else "",
            "di": str(di) if di else "",
            "eta": eta if isinstance(eta, datetime) else None,
            "etd": etd if isinstance(etd, datetime) else None,
            "atd": atd if isinstance(atd, datetime) else None,
            "weight": weight,
            "batch": str(batch) if batch else "",
            "device": str(device) if device else "",
            "waybill": str(waybill) if waybill else "",
            "key": key,
        })

    # 按ETA排序，附加序号
    boxes.sort(key=lambda x: x["eta"] or datetime(2099, 1, 1))
    for i, b in enumerate(boxes):
        b["seq"] = i + 1

    # 缓存
    if isinstance(wb_or_path, _OpenpyxlWorkbookProxy):
        _cache[cache_key] = boxes

    return boxes


def _find_summary_data_start(ws):
    """找汇总表的数据起始行"""
    for row in range(1, 20):
        for col in range(1, 20):
            val = ws.cell(row=row, column=col).value
            if val and "箱号" in str(val):
                return row + 1
    return 6


# ══════════════════════════════════════════════════════
#  读取订单数据
# ══════════════════════════════════════════════════════

def get_current_orders(wb_or_path):
    """
    读取「配箱公式」sheet，返回当前已导入的订单及配箱结果
    支持 wb 对象或文件路径
    """
    if isinstance(wb_or_path, str):
        wb = load_allocation_workbook(wb_or_path)
        try:
            return get_current_orders(wb)
        finally:
            wb.close()

    # 检查缓存
    if isinstance(wb_or_path, _OpenpyxlWorkbookProxy):
        cache_key = _get_cache_key(wb_or_path._file_path) + "|orders"
        if cache_key in _cache:
            return _cache[cache_key]

    ws = wb_or_path["配箱公式"]
    orders = []

    # 如果是 xlwings proxy，用批量读取提速
    if isinstance(ws, _XlwingsSheetProxy):
        orders = _read_orders_xlwings(ws)
    else:
        orders = _read_orders_openpyxl(ws)

    # 缓存
    if isinstance(wb_or_path, _OpenpyxlWorkbookProxy):
        _cache[cache_key] = orders

    return orders


def _read_orders_xlwings(ws):
    """xlwings 批量读取订单"""
    # 一次性读取所需列的范围（Z~BJ, H, J, M, S）
    # 用 cell 逐行读取（数据已在内存中，速度快）
    orders = []
    for row in range(2, ws.max_row + 1):
        order_no = ws.cell(row=row, column=26).value  # Z列
        if not order_no:
            continue

        order_str = str(int(order_no)) if isinstance(order_no, (int, float)) else str(order_no).strip()

        sales_org = ws.cell(row=row, column=29).value  # AC列
        material = ws.cell(row=row, column=30).value  # AD列
        brand     = ws.cell(row=row, column=8).value    # H列
        qty       = ws.cell(row=row, column=35).value  # AI列
        box_no    = ws.cell(row=row, column=10).value   # J列
        di        = ws.cell(row=row, column=19).value    # S列
        device    = ws.cell(row=row, column=13).value   # M列
        batch     = ws.cell(row=row, column=62).value   # BJ列
        ship_point= ws.cell(row=row, column=46).value   # AT列

        status = _determine_status(box_no, brand)
        orders.append(_build_order_dict(row, order_str, sales_org, material,
                                         brand, qty, box_no, di, device,
                                         batch, ship_point, status))
    return orders


def _read_orders_openpyxl(ws):
    """openpyxl 逐行读取订单"""
    orders = []
    for row in range(2, ws.max_row + 1):
        order_no = ws.cell(row=row, column=26).value  # Z列
        if not order_no:
            continue

        order_str = str(int(order_no)) if isinstance(order_no, (int, float)) else str(order_no).strip()

        sales_org = ws.cell(row=row, column=29).value
        material = ws.cell(row=row, column=30).value
        brand     = ws.cell(row=row, column=8).value
        qty       = ws.cell(row=row, column=35).value
        box_no    = ws.cell(row=row, column=10).value
        di        = ws.cell(row=row, column=19).value
        device    = ws.cell(row=row, column=13).value
        batch     = ws.cell(row=row, column=62).value
        ship_point= ws.cell(row=row, column=46).value

        status = _determine_status(box_no, brand)
        orders.append(_build_order_dict(row, order_str, sales_org, material,
                                         brand, qty, box_no, di, device,
                                         batch, ship_point, status))
    return orders


def _determine_status(box_no, brand):
    """判断订单配箱状态"""
    if box_no and str(box_no).strip() and str(box_no).strip() not in ("#N/A", "/"):
        return "已配"
    elif brand and str(brand).strip() != "#N/A":
        return "待配"
    else:
        return "无库存"


def _build_order_dict(row, order_str, sales_org, material, brand, qty,
                      box_no, di, device, batch, ship_point, status):
    """构建订单字典"""
    return {
        "row": row,
        "order_no": order_str,
        "sales_org": str(sales_org) if sales_org else "",
        "material": str(material) if material else "",
        "brand": str(brand) if brand else "",
        "qty": qty,
        "box_no": str(box_no).strip() if box_no and str(box_no).strip() not in ("#N/A", "/") else None,
        "di": str(di).strip() if di and str(di).strip() != "#N/A" else "",
        "batch": str(batch).strip() if batch and str(batch).strip() != "#N/A" else "",
        "device": str(device).strip() if device and str(device).strip() != "#N/A" else "",
        "ship_point": str(ship_point).strip() if ship_point and str(ship_point).strip() != "#N/A" else "",
        "status": status,
    }


# ══════════════════════════════════════════════════════
#  xlwings 重算
# ══════════════════════════════════════════════════════

def refresh_calculation(file_path):
    """
    用 xlwings 打开文件，强制 Excel 重算公式，然后保存关闭
    如果 xlwings 不可用则跳过（openpyxl 无法触发公式重算）
    返回 True 表示重算成功，False 表示跳过
    """
    if not has_xlwings():
        return False

    import xlwings as xw
    app = None
    book = None
    try:
        app = xw.App(visible=False)
        app.calculation = "automatic"
        book = app.books.open(file_path)
        app.calculate()
        book.save()
        book.close()
        book = None
        app.quit()
        app = None
        _invalidate_cache(file_path)
        return True
    except Exception:
        return False
    finally:
        if book:
            try:
                book.close()
            except Exception:
                pass
        if app:
            try:
                app.quit()
            except Exception:
                pass


# ══════════════════════════════════════════════════════
#  配箱计算逻辑
# ══════════════════════════════════════════════════════

def allocate_boxes(orders, boxes):
    """
    配箱计算：按行顺序，为每个订单分配可用箱子
    """
    pool = {}
    for b in boxes:
        k = b["key"]
        if k not in pool:
            pool[k] = []
        pool[k].append(b)

    allocated = []
    used_boxes = set()

    for order in orders:
        if order["status"] == "指定" and order.get("_manual_box"):
            allocated.append(order)
            used_boxes.add(order["_manual_box"])
            continue

        key = f"{order['qty']}|{order['sales_org']}|{order['brand']}"
        if key not in pool or not pool[key]:
            order["status"] = "无库存"
            allocated.append(order)
            continue

        box = pool[key].pop(0)
        order["box_no"] = box["box_no"]
        order["status"] = "已配"
        order["_allocated_from"] = box
        used_boxes.add(box["box_no"])
        allocated.append(order)

    remaining = [b for b in boxes if b["box_no"] not in used_boxes]
    return allocated, remaining


def manual_assign_box(orders, boxes, order_rows, box_no):
    """
    手动指定箱号
    """
    box = None
    for b in boxes:
        if b["box_no"] == box_no:
            box = b
            break

    if not box:
        return None, f"箱号 {box_no} 不在可用库存池中"

    for order in orders:
        if order["row"] in order_rows:
            order["box_no"] = box_no
            order["status"] = "指定"
            order["_manual_box"] = box_no

    return orders, None


def manual_assign_di(orders, boxes, order_rows, di):
    """
    手动指定DI：该DI下的箱子按ETA先到先配
    """
    di_boxes = [b for b in boxes if b["di"] == di]
    di_boxes.sort(key=lambda x: x["eta"] or datetime(2099, 1, 1))

    if not di_boxes:
        return None, f"DI {di} 下没有可用的箱子"

    order_idx = 0
    for order in orders:
        if order["row"] in order_rows:
            if order_idx < len(di_boxes):
                box = di_boxes[order_idx]
                order["box_no"] = box["box_no"]
                order["status"] = "指定"
                order["_manual_box"] = box["box_no"]
                order_idx += 1
            else:
                order["status"] = "无库存"

    return orders, None


# ══════════════════════════════════════════════════════
#  ERP订单导入
# ══════════════════════════════════════════════════════

def import_erp_orders(erp_file_path, allocation_file_path, mapping=None):
    """
    导入ERP订单到配箱表
    写入后尝试用 xlwings 重算公式
    """
    if mapping is None:
        cfg = load_mapping_config()
        mapping = cfg.get("erp_mapping", DEFAULT_ERP_MAPPING)

    # 读取ERP文件
    erp_wb = load_workbook(erp_file_path, data_only=True)
    erp_ws = erp_wb.active

    header_row = _find_erp_header(erp_ws)
    if header_row == 0:
        return 0, "无法识别ERP文件表头，请检查文件格式"

    erp_col_map = {}
    for col in range(1, erp_ws.max_column + 1):
        val = erp_ws.cell(row=header_row, column=col).value
        if val:
            erp_col_map[str(val).strip()] = col

    # 打开配箱表（带公式版本，用于写入）
    alloc_wb = load_workbook(allocation_file_path)
    alloc_ws = alloc_wb["配箱公式"]

    next_row = _find_next_data_row(alloc_ws)

    success_count = 0
    for data_row in range(header_row + 1, erp_ws.max_row + 1):
        order_no_col = erp_col_map.get("客户订单号")
        if not order_no_col:
            break
        order_val = erp_ws.cell(row=data_row, column=order_no_col).value
        if not order_val:
            continue

        for erp_col_name, alloc_col in mapping.items():
            erp_col_idx = erp_col_map.get(erp_col_name)
            if not erp_col_idx:
                continue
            val = erp_ws.cell(row=data_row, column=erp_col_idx).value
            if val is None:
                continue

            alloc_col_idx = column_index_from_string(alloc_col)
            alloc_ws.cell(row=next_row, column=alloc_col_idx).value = val

        material = erp_ws.cell(row=data_row, column=erp_col_map.get("物料名称", 0)).value
        if material:
            alloc_ws.cell(row=next_row, column=1).value = material

        next_row += 1
        success_count += 1

    alloc_wb.save(allocation_file_path)
    erp_wb.close()

    # 尝试重算
    _invalidate_cache(allocation_file_path)
    refresh_calculation(allocation_file_path)

    return success_count, None


def _find_erp_header(ws):
    """找ERP文件的表头行"""
    for row in range(1, 10):
        for col in range(1, 20):
            val = ws.cell(row=row, column=col).value
            if val and ("客户订单号" in str(val) or "订单号" in str(val)):
                return row
    return 0


def _find_next_data_row(ws):
    """找配箱公式sheet的下一个空行"""
    for row in range(DATA_START_ROW, ws.max_row + 2):
        val = ws.cell(row=row, column=26).value
        if not val:
            return row
    return ws.max_row + 1


# ══════════════════════════════════════════════════════
#  ASN导入 + 运单号补录
# ══════════════════════════════════════════════════════

def import_asn(asn_file_path, allocation_file_path):
    """
    导入ASN数据到配箱表
    写入后尝试用 xlwings 重算公式
    """
    asn_wb = load_workbook(asn_file_path, data_only=True)
    asn_ws = asn_wb.active

    alloc_wb = load_workbook(allocation_file_path)
    alloc_ws = alloc_wb["ASN"]

    # 清空ASN sheet（保留表头）
    for row in range(2, alloc_ws.max_row + 1):
        for col in range(1, alloc_ws.max_column + 1):
            alloc_ws.cell(row=row, column=col).value = None

    # 复制ASN数据
    for row in range(1, asn_ws.max_row + 1):
        for col in range(1, asn_ws.max_column + 1):
            val = asn_ws.cell(row=row, column=col).value
            alloc_ws.cell(row=row, column=col).value = val

    alloc_wb.save(allocation_file_path)
    asn_wb.close()

    # 尝试重算
    _invalidate_cache(allocation_file_path)
    refresh_calculation(allocation_file_path)

    return True, None


def supplement_waybill(allocation_file_path, box_no, waybill_no):
    """补录运单号"""
    wb = load_workbook(allocation_file_path)
    ws = wb["ASN"]

    found = False
    for row in range(1, ws.max_row + 1):
        bn = ws.cell(row=row, column=1).value
        if bn and str(bn).strip() == box_no:
            ws.cell(row=row, column=2).value = waybill_no
            found = True
            break

    wb.save(allocation_file_path)
    _invalidate_cache(allocation_file_path)
    return found, None


# ══════════════════════════════════════════════════════
#  导出
# ══════════════════════════════════════════════════════

def export_allocation_list(orders, output_path):
    """导出配箱清单"""
    wb = Workbook()
    ws = wb.active
    ws.title = "配箱清单"

    ws["A1"] = "箱号"
    ws["B1"] = "PO"
    ws["C1"] = "客户订单号"
    ws["D1"] = "销售组织"
    ws["E1"] = "牌号"
    ws["F1"] = "数量"

    row = 2
    for order in orders:
        if order["box_no"]:
            ws.cell(row=row, column=1).value = order["box_no"]
            ws.cell(row=row, column=2).value = order.get("po", "")
            ws.cell(row=row, column=3).value = order["order_no"]
            ws.cell(row=row, column=4).value = order["sales_org"]
            ws.cell(row=row, column=5).value = order["brand"]
            ws.cell(row=row, column=6).value = order["qty"]
            row += 1

    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════
#  工具函数
# ══════════════════════════════════════════════════════

def format_eta(dt):
    """格式化ETA/ETD为字符串"""
    if isinstance(dt, datetime):
        return dt.strftime("%Y/%m/%d")
    return str(dt) if dt else ""


def get_hub_from_di(di_str):
    """从DI前缀获取Hub"""
    if not di_str or len(di_str) < 2:
        return ""
    prefix = di_str[:2].upper()
    hub_map = {
        "SH": "上海Hub",
        "HP": "黄埔Hub",
        "QD": "青岛Hub",
        "NB": "宁波Hub",
        "ST": "汕头Hub",
        "XG": "新港Hub",
        "XM": "厦门Hub",
        "NS": "南沙Hub",
        "WF": "潍坊Hub",
        "NJ": "南京Hub",
        "TC": "太仓Hub",
        "SQ": "宿迁Hub",
        "CQ": "重庆Hub",
        "WH": "武汉Hub",
        "TZ": "台州Hub",
        "HF": "合肥Hub",
        "FZ": "福州Hub",
    }
    return hub_map.get(prefix, "")
